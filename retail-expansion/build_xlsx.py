#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Assemble le chiffrier des points de vente candidats de Lasclay."""

import json, math, re, unicodedata, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
J = lambda n: json.load(open(os.path.join(BASE, n), encoding='utf-8'))

# ---------------------------------------------------------------- reseau actuel
ACTUELS = [
    ("Lasclay – Boutique et Atelier", "Boutique et atelier de la marque", "Québec", "QC",
     "254 boul. des Capucins, 2e étage", "G1J 3R4", "", "", "https://lasclay.com", "Tous + exclusivités + prototypes"),
    ("Les Défricheuses", "Boutique écoresponsable / créatrices d'ici", "Montréal", "QC",
     "1319 rue Beaubien E", "H2G 1K7", "514-374-5325", "", "https://lesdefricheuses.ca", "Tous les produits"),
    ("L'Axe du Malt", "Boutique de bière et de brassage", "Québec", "QC",
     "601 rue D'Aiguillon", "G1R 3H3", "", "", "", "Sacs à lunch, glacières, manchons isothermes"),
    ("Vert métal", "Boutique écoresponsable", "L'Ancienne-Lorette", "QC",
     "2051 rue Saint-Jean-Baptiste", "G2E 1R6", "", "", "https://www.vertmetal.com", "Sacs à lunch, sacs à vin, coussins, cuisine, manchons"),
    ("Agriculture du Coin", "Jardinerie et agriculture urbaine", "Montréal", "QC",
     "235 av. Laurier O", "H2T 2N9", "514-273-4332", "", "https://agricultureducoin.ca", "Sacs à lunch"),
    ("Parc national des Îles-de-Boucherville (Sépaq)", "Boutique de parc national", "Boucherville", "QC",
     "55 Île Sainte-Marguerite", "J4B 5J6", "450-928-5088", "", "https://www.sepaq.com/pq/bou", "Selon la saison"),
    ("Belle et Rebelle", "Boutique de créateurs québécois", "Sherbrooke", "QC",
     "90 rue Wellington N", "J1H 5B8", "819-791-2299", "", "https://www.belleetrebellesherbrooke.ca", "Mitaines"),
    ("Chez Georges-Émile", "Boutique cadeaux et déco", "Rimouski", "QC",
     "92 rue St-Germain O", "G5L 4B5", "418-725-8533", "", "", "Sacs à lunch"),
    ("Librairie Alpha", "Librairie indépendante", "Gaspé", "QC",
     "168 rue de la Reine", "G4X 1T4", "418-368-5514", "", "https://alpha.leslibraires.ca", "Sacs à lunch"),
    ("Local Refillery", "Refillery / zéro déchet", "Courtenay", "BC",
     "420 Fitzgerald Ave", "V9N 7N2", "250-871-6600", "", "https://www.localrefillery.com", "Sacs à lunch"),
]
# Zones ou un detaillant Lasclay est deja en place. La Rive-Sud n'y figure pas:
# la boutique du parc des Iles-de-Boucherville est saisonniere et ne porte qu'une
# partie de la gamme, donc un detaillant a l'annee reste a trouver.
ZONES_COUVERTES = {
    "Quebec - centre et Limoilou", "Montreal - centre et Plateau",
    "Sherbrooke", "Rimouski", "Gaspe et Perce", "Comox Valley",
}

# ------------------------------------------------------------------- referentiel
SCORE_TYPE = {
    "Cadeaux / artisans": 40, "Artisans quebecois": 40, "Artisans canadiens": 40,
    "Artisans atlantiques": 40, "Artisans et designers quebecois": 40,
    "Cadeaux / produits quebecois": 40, "Designers canadiens": 40,
    "Metiers d'art": 36, "Artisans / metiers d'art": 36, "Artisans": 34,
    "Artisans / cadeaux": 38, "Artisans / galerie": 32, "Galerie / artisans": 30,
    "Eco / zero dechet": 35, "Magasin general ecoresponsable": 35,
    "Magasin general eco": 35, "Eco / boutique": 34, "Eco / artisans": 38,
    "Eco / produits quebecois": 38, "Boutique eco": 32, "Eco / plein air": 38,
    "Plein air": 35, "Jardinerie / horticulture": 32,
    "Epicerie fine / terroir": 30, "Terroir / cadeaux quebecois": 38,
    "Terroir / epicerie fine": 30, "Terroir / produits regionaux": 36,
    "Fromagerie / terroir": 28, "Kiosque fermier / terroir": 28,
    "Ornithologie / nature": 40, "Boutique de musee": 34,
    "Librairie / artisans": 34, "Librairie independante": 26,
    "Alimentation sante / bio": 24, "Epicerie bio / vrac": 24, "Epicerie bio / eco": 24,
    "Epicerie zero dechet": 28, "Epicerie sante / eco": 24, "Epicerie / vrac": 24,
    "Epicerie vrac": 24, "Marche / eco": 26, "Eco / marche local": 30,
    "Cafe-boutique eco": 30, "Eco / cosmetiques": 20, "Eco / apothicaire": 22,
    "Eco / sante": 20, "Eco / bien-etre": 20, "Cadeaux / eco": 34,
    "Sante naturelle / boutique": 20, "Coop artisans": 36, "Cooperative artisans": 36,
    "Artisanat / materiaux creatifs": 18, "Friperie / eco": 16, "Eco / friperie": 18,
    "Fleuriste / jardin": 18, "Plein air / sport": 22, "Chasse et peche": 28,
    "Vin / spiritueux": 14, "Boissons / brassage": 28, "Maroquinerie / sacs": 24,
    "Deco / maison": 26, "Articles de maison": 26,
    "Eco / grande boutique verte": 30, "Boutique / cadeaux": 30,
}
# Le marche compte plus que l'etiquette: on note la population reelle de la zone.
# Consigne: privilegier les villes de 50 000 habitants et plus.
def score_population(p):
    if p >= 500000: return 30
    if p >= 150000: return 26
    if p >= 50000:  return 22
    if p >= 20000:  return 10
    if p >= 5000:   return 5
    return 2

def tranche_population(p):
    if p >= 500000: return "500 000 et plus"
    if p >= 150000: return "150 000 a 500 000"
    if p >= 50000:  return "50 000 a 150 000"
    if p >= 20000:  return "20 000 a 50 000"
    if p >= 5000:   return "5 000 a 20 000"
    return "moins de 5 000"

# chaines et grandes surfaces a exclure (ne correspondent pas au modele consignation)
CHAINES = [
    "dollarama", "dollar tree", "hallmark", "carlton card", "things engraved", "winners",
    "walmart", "canadian tire", "home hardware", "rona", "home depot", "lowe", "costco",
    "sport chek", "atmosphere", "sportium", "decathlon", "mountain equipment", "mec ",
    "bulk barn", "nutrition house", "gnc", "popeye", "vitamin shop", "sephora", "lush",
    "the body shop", "bath & body", "yves rocher", "saje", "davidstea", "second cup",
    "starbucks", "tim horton", "mcdonald", "subway", "shoppers drug", "pharmaprix",
    "jean coutu", "uniprix", "familiprix", "metro ", "iga ", "provigo", "maxi ", "loblaw",
    "sobeys", "safeway", "save-on-food", "no frills", "super c", "food basics",
    "indigo", "chapters", "coles", "renaud-bray", "archambault", "toys r us",
    "spencer", "claire's", "ardene", "la senza", "dynamite", "garage", "reitmans",
    "value village", "goodwill", "salvation army", "armee du salut", "talize",
    "canex", "hudson's bay", "la baie", "simons", "holt renfrew", "nordstrom",
    "michaels", "deserres", "omer deserres", "staples", "bureau en gros",
    "sherwin", "benjamin moore", "pet valu", "mondou", "petsmart",
    # « gift shop » et « souvenir » ne sont pas disqualifiants: ce sont des
    # descriptifs que portent de vraies boutiques d'artisans. Seuls les points
    # de vente captifs le sont.
    "duty free", "hors taxe", "airport", "aeroport",
    # succursales et grandes surfaces d'alcool: achat par centrale, aucun interet
    "liquor", "lcbo", "saq ", "societe des alcools", "beer store", "brewers retail",
    "wine rack", "cold beer", "spirits", "alcool", "wine & beyond", "ace discount",
    "cannabis", "vape", "smoke shop", "tobacco",
    # autres enseignes sans pertinence
    "petro", "esso", "shell", "circle k", "couche-tard", "7-eleven", "ultramar",
    "electrolux", "vacuum", "aspirateur", "kirby", "rent-a-center", "cash",
    # etablissements qui ne sont pas des commerces de detail
    "karate", "dojo", "taekwondo", "judo", "jiu-jitsu", "crossfit", "gym ",
    "fitness", "arena", "curling", "bowling", "golf club", "country club",
    "distribution", "distributeur", "wholesale", "grossiste", "manufactur",
    "warehouse inc", "entrepot", "imports inc", "holdings", "vileda",
    "cleaning products", "supply co. ltd", "industries",
    "u-haul", "storage", "self storage", "car wash", "lave-auto",
]

def strip_acc(s):
    if not s: return ""
    return "".join(c for c in unicodedata.normalize("NFD", str(s))
                   if unicodedata.category(c) != "Mn")

def norm(s):
    # les noms de zone utilisent des tirets cadratins: on ecrase toute
    # ponctuation, sinon la comparaison avec ZONES_COUVERTES echoue en silence.
    return re.sub(r"[^a-z0-9]+", "", strip_acc(s).lower())

def cle_zone(s):
    return norm(s)

def clean_addr(a):
    if not a: return ""
    a = re.sub(r"\s+", " ", str(a)).strip(" ,;|-")
    # coupe le bruit d'horaires ramasse par le scraper
    a = re.split(r"(?i)\b(?:lundi|mardi|mercredi|jeudi|vendredi|samedi|dimanche|monday|tuesday|hours|heures|ouvert|open)\b", a)[0]
    a = re.sub(r"\d{1,2}\s?h\s?(a|à|-)\s?\d{1,2}\s?h", "", a)
    return re.sub(r"\s+", " ", a).strip(" ,;|-")[:90]

MAUVAIS_MAIL = re.compile(
    r"\.(png|jpe?g|gif|svg|webp|avif|heic|bmp|tiff?|css|js|woff2?)$|^[0-9a-f-]{24,}@", re.I)

def clean_mail(m):
    m = (m or "").strip().lower()
    # le scraper ramasse parfois un nom de fichier d'image comme adresse
    return "" if (not m or MAUVAIS_MAIL.search(m) or len(m) > 60) else m

def clean_tel(t):
    if not t: return ""
    d = re.sub(r"\D", "", str(t))
    if len(d) == 11 and d.startswith("1"): d = d[1:]
    if len(d) != 10: return str(t)[:24]
    return f"{d[0:3]}-{d[3:6]}-{d[6:10]}"

def haversine(a, b, c, d):
    R = 6371.0
    p1, p2 = math.radians(a), math.radians(c)
    dp, dl = math.radians(c - a), math.radians(d - b)
    h = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * R * math.asin(math.sqrt(h))


# ------------------------------------------------------------------- jugement
# Les etiquettes OpenStreetMap sont larges. Ces regles font le tri qu'un humain
# ferait en lisant la liste: elles ecartent ce qui ne peut pas porter Lasclay et
# remontent ce qui a les signaux d'une boutique independante a recit.

# archetypes ecartes d'office: aucun lien plausible avec la gamme
ARCHETYPES_EXCLUS = {
    "Artisanat / materiaux creatifs",   # fournitures de bricolage, pas de vente d'objets finis
    "Friperie / eco",                   # seconde main: ne prend pas de neuf en consignation
    "Eco / friperie",
    "Maroquinerie / sacs",              # concurrence directe sur le sac
}

# archetypes gardes seulement si le nom porte un signal de specialite
ARCHETYPES_CONDITIONNELS = {
    "Vin / spiritueux": ("cave", "vignoble", "vigneron", "winery", "cellar", "microbrasserie",
                         "brasserie", "brewing", "brewery", "malt", "cidrerie", "hydromel",
                         "distillerie", "distillery", "terroir", "importation privee"),
    "Plein air / sport": ("plein air", "outdoor", "outfitter", "adventure", "aventure",
                          "mountain", "montagne", "trail", "camp", "canoe", "kayak",
                          "randonnee", "hiking", "ski", "nature", "expedition"),
    "Fleuriste / jardin": ("jardin", "garden", "centre", "center", "serre", "pepiniere",
                           "nursery", "greenhouse", "horticole"),
    # Deco et articles de maison restent sans condition de nom: Chez
    # Georges-Emile, deja detaillant, est precisement une boutique cadeaux et
    # deco. Coussins, mitaines de four, sous-plats et couvertures y ont leur
    # place. Leur pointage plus bas suffit a les classer apres les autres.
}

# mots du nom qui trahissent une boutique independante a recit
SIGNAUX_POSITIFS = ("artisan", "artisans", "artisanat", "local", "locaux", "made in",
                    "coop", "cooperative", "marche", "market", "terroir", "ferme", "farm",
                    "atelier", "nature", "jardin", "garden", "vrac", "bulk", "refill",
                    "eco", "ecolo", "vert", "green", "zero dechet", "zero waste",
                    "general store", "magasin general", "mercantile", "trading post",
                    "handmade", "fait main", "createur", "createurs", "maker", "craft",
                    "boutique", "du coin", "village", "heritage", "quebec", "gaspesie",
                    "charlevoix", "acadie", "maritimes")

# mots du nom qui disent grande surface, liquidation ou franchise
SIGNAUX_NEGATIFS = ("outlet", "liquidation", "clearance", "discount", "depot", "mart",
                    "superstore", "megastore", "warehouse", "wholesale", "franchise",
                    "express", "express inc", "self serve", "vending")

def jugement(r):
    """Renvoie (bonus, motif_de_rejet). Un motif non vide ecarte le candidat."""
    nom = strip_acc(r["nom"]).lower()
    typ = r["type"]

    if typ in ARCHETYPES_EXCLUS:
        return 0, f"archetype ecarte ({typ})"

    cond = ARCHETYPES_CONDITIONNELS.get(typ)
    if cond and not any(m in nom for m in cond):
        return 0, f"{typ} sans signal de specialite dans le nom"

    if any(m in nom for m in SIGNAUX_NEGATIFS):
        return 0, "nom de grande surface ou de liquidation"

    bonus = 0
    touches = [m for m in SIGNAUX_POSITIFS if m in nom]
    if touches:
        bonus += min(12, 4 * len(touches))
    return bonus, ""

def moyens_de_contact(r):
    m = []
    if r.get("tel"): m.append("téléphone")
    if r.get("courriel"): m.append("courriel")
    if r.get("fb") or r.get("ig"): m.append("réseaux sociaux")
    if r.get("web"): m.append("site web")
    return ", ".join(m)

# ------------------------------------------------------------------------ zones
zones = J("zones.json")
ville2zone = {}
for z in zones:
    for v in z["villes"]:
        ville2zone.setdefault((z["prov"], norm(v)), z)
    ville2zone.setdefault((z["prov"], norm(z["ancre"])), z)

def zone_la_plus_proche(lat, lon, seuil=90.0):
    best, bd = None, 1e9
    for z in zones:
        if not z.get("lat"): continue
        d = haversine(lat, lon, z["lat"], z["lon"])
        if d < bd: best, bd = z, d
    if best and bd <= seuil:
        return best, f"a {bd:.0f} km de {best['ancre']}"
    return None, ""

def trouver_zone(prov, ville, lat=None, lon=None):
    z = ville2zone.get((prov, norm(ville)))
    if z: return z, "ville"
    if lat and lon:
        best, bd = None, 1e9
        for zz in zones:
            if not zz.get("lat"): continue
            d = haversine(float(lat), float(lon), zz["lat"], zz["lon"])
            if d < bd: best, bd = zz, d
        if best and bd <= 60:
            return best, f"proximite ({bd:.0f} km)"
    return None, ""

# -------------------------------------------------------------------- chargement
rows, vus = [], set()

zone_par_nom = {z["zone"]: z for z in zones}

def ajouter(nom, typ, ville, prov, adresse, cp, tel, mail, web, source,
            lat=None, lon=None, zone=None, fb="", ig=""):
    if not nom or len(nom) < 2: return
    n = norm(nom)
    if any(c in strip_acc(nom).lower() for c in CHAINES): return
    # les coordonnees priment: chaque commerce va a la zone-ancre la plus proche.
    # a defaut, on utilise la ville inscrite, puis la zone du rayon de moisson.
    z, moyen = (None, "")
    if lat and lon:
        z, moyen = zone_la_plus_proche(float(lat), float(lon))
    if not z and ville:
        z, moyen = ville2zone.get((prov, norm(ville))), "ville"
    if not z and zone and zone in zone_par_nom:
        z, moyen = zone_par_nom[zone], "rayon de la ville-ancre"
    if not z:
        z, moyen = trouver_zone(prov, ville, lat, lon)
    if not z: return
    # dedoublonnage global: les rayons de moisson se chevauchent
    cle = (n, round(float(lat), 3), round(float(lon), 3)) if lat and lon else (n, z["zone"])
    if cle in vus: return
    vus.add(cle)
    rows.append(dict(nom=nom.strip(), type=typ, zone=z["zone"], ancre=z["ancre"],
                     palier=z["palier"], pop=z.get("pop", 0),
                     ville=(ville or "").strip(), prov=z["prov"],
                     adresse=clean_addr(adresse), cp=(cp or "").strip().upper(),
                     tel=clean_tel(tel), courriel=clean_mail(mail),
                     web=(web or "").strip(), fb=(fb or "").strip(), ig=(ig or "").strip(),
                     source=source, rattachement=moyen))

for fichier in ("candidats_enrichis.json", "candidats_zones_manquantes_enrichis.json"):
    try:
        lot = J(fichier)
    except Exception:
        continue
    for c in lot:
        ajouter(c.get("nom"), strip_acc(c.get("type", "")), c.get("ville"), c.get("prov"),
                c.get("adresse"), c.get("cp"), c.get("tel"), c.get("courriel"),
                c.get("web"), "Recherche web ciblee", fb=c.get("fb", ""), ig=c.get("ig", ""))

try:
    osm = J("osm_brut.json")
except Exception:
    osm = []
for c in osm:
    ajouter(c.get("nom"), strip_acc(c.get("type", "")), c.get("ville"), c.get("prov"),
            c.get("adresse"), c.get("cp"), c.get("tel"), c.get("courriel"),
            c.get("web"), "OpenStreetMap", c.get("lat"), c.get("lon"), c.get("zone"),
            c.get("fb"), c.get("ig"))

# --- coordonnees trouvees sur les sites des boutiques -------------------------
par_web = {}
for r in rows:
    if r["web"]:
        par_web.setdefault(r["web"].rstrip("/").lower(), []).append(r)
sites_morts = set()
try:
    for e in J("a_scraper_enrichi.json"):
        if e.get("_mort"): sites_morts.add((e.get("web") or "").rstrip("/").lower())
        for r in par_web.get((e.get("web") or "").rstrip("/").lower(), []):
            if not r["tel"] and e.get("tel"): r["tel"] = clean_tel(e["tel"])
            if not r["courriel"] and e.get("courriel"): r["courriel"] = clean_mail(e["courriel"])
            if not r["adresse"] and e.get("adresse"): r["adresse"] = clean_addr(e["adresse"])
            if not r["cp"] and e.get("cp"): r["cp"] = e["cp"]
except Exception:
    pass

# --- adresses qui ne sont pas celles du commerce --------------------------------
# Le scraper ramasse ce qui ressemble a un courriel dans la page, y compris
# l'adresse du developpeur dans le pied de page ou celle de l'auteur d'une
# police dans un commentaire CSS. Signature: une meme adresse rattachee a trois
# commerces de noms differents. On la retire, la fiche garde son telephone.
noms_par_mail = {}
for r in rows:
    if r["courriel"]:
        noms_par_mail.setdefault(r["courriel"], set()).add(norm(r["nom"]))
etrangeres = {m for m, n in noms_par_mail.items() if len(n) >= 3}
if etrangeres:
    print(f"Adresses retirees (rattachees a 3 commerces ou plus) : {len(etrangeres)}")
    for r in rows:
        if r["courriel"] in etrangeres:
            r["courriel"] = ""

# --- tri au jugement -----------------------------------------------------------
# 1. une enseigne presente dans trois zones ou plus est une chaine: elle achete
#    par centrale et ne fera pas de consignation, quel que soit son nom.
zones_par_enseigne = {}
for r in rows:
    zones_par_enseigne.setdefault(norm(r["nom"]), set()).add(r["zone"])

garde, rejets = [], {}
for r in rows:
    if len(zones_par_enseigne[norm(r["nom"])]) >= 4:
        rejets["enseigne presente dans 4 zones ou plus"] = rejets.get("enseigne presente dans 4 zones ou plus", 0) + 1
        continue
    bonus, motif = jugement(r)
    if motif:
        rejets[motif.split(" (")[0]] = rejets.get(motif.split(" (")[0], 0) + 1
        continue
    # 2. sans moyen de joindre le commerce, la fiche est inutilisable. Un site
    #    qui repond compte: on y trouve un formulaire. Un domaine mort, non.
    site_vivant = bool(r["web"]) and r["web"].rstrip("/").lower() not in sites_morts
    if not (r["tel"] or r["courriel"] or r["fb"] or r["ig"] or site_vivant):
        rejets["aucun moyen de contact"] = rejets.get("aucun moyen de contact", 0) + 1
        continue
    r["bonus"] = bonus
    garde.append(r)
rows = garde
print("Ecartes au tri :")
for k, v in sorted(rejets.items(), key=lambda x: -x[1]):
    print(f"  {v:>6}  {k}")

# ---------------------------------------------------------------------- pointage
for r in rows:
    s = SCORE_TYPE.get(r["type"], 22)
    s += score_population(r["pop"])
    s += 7 if r["tel"] else 0
    s += 7 if r["courriel"] else 0
    s += 6 if r["adresse"] else 0
    s += 5 if r["web"] else 0
    if r["source"] == "Recherche web ciblee": s += 10  # archetype valide a la lecture
    s += r.get("bonus", 0)                              # signaux du nom
    if r["tel"] and r["courriel"]: s += 5                # joignable de deux facons
    if cle_zone(r["zone"]) in {cle_zone(x) for x in ZONES_COUVERTES}:
        s -= 20
        r["statut_zone"] = "Zone deja couverte"
    else:
        s += 15
        r["statut_zone"] = "Zone a ouvrir"
    r["score"] = s

rows.sort(key=lambda r: (-r["score"], r["zone"], r["nom"]))

# rang a l'interieur de la zone (regle d'exclusivite: on approche le rang 1 d'abord)
par_zone = {}
for r in rows:
    par_zone.setdefault(r["zone"], []).append(r)
    r["rang"] = len(par_zone[r["zone"]])
    r["priorite"] = ("A - approcher en premier" if r["rang"] == 1 and r["statut_zone"] == "Zone a ouvrir"
                     else "B - releve" if r["rang"] <= 3 and r["statut_zone"] == "Zone a ouvrir"
                     else "C - reserve")

# Cibles: 500 candidats au Quebec, 1000 dans le reste du Canada.
# Deux temps. D'abord le rang 1 de chaque zone, pour que la regle de l'heure de
# route tienne partout. Ensuite on approfondit, en servant les zones de 50 000
# habitants et plus avant les autres.
CIBLES = {"QC": 500, "ROC": 1000}
GROS = 50000
retenus = []
for groupe, cible in CIBLES.items():
    pool = [r for r in rows if (r["prov"] == "QC") == (groupe == "QC")]
    pris = sorted([r for r in pool if r["rang"] == 1], key=lambda r: -r["score"])[:cible]
    deja = {id(r) for r in pris}
    reste = sorted([r for r in pool if id(r) not in deja],
                   key=lambda r: (r["pop"] < GROS, r["rang"], -r["score"]))
    pris += reste[:max(0, cible - len(pris))]
    retenus += pris
retenus.sort(key=lambda r: (r["prov"], r["zone"], r["rang"]))

# coordonnees completees a partir des sites web (scrape_contacts.js)
cle_contact = {(norm(r["nom"]), r["zone"]): r for r in retenus}
try:
    for e in J("selection_enrichie.json"):
        c = cle_contact.get((norm(e.get("nom")), e.get("zone")))
        if not c: continue
        for champ in ("tel", "courriel", "adresse", "cp"):
            if not c.get(champ) and e.get(champ):
                c[champ] = clean_tel(e[champ]) if champ == "tel" else (
                    clean_addr(e[champ]) if champ == "adresse" else e[champ])
except Exception:
    pass

json.dump([{k: v for k, v in r.items()} for r in retenus],
          open(os.path.join(BASE, "selection.json"), "w", encoding="utf-8"),
          ensure_ascii=False, indent=1)


# Les libelles internes sont sans accents (comparaisons robustes). On les
# retablit au moment d'ecrire, parce que le chiffrier se lit en francais.
ACCENTS = {
    "Eco / zero dechet": "Éco / zéro déchet",
    "Eco / artisans": "Éco / artisans",
    "Eco / produits quebecois": "Éco / produits québécois",
    "Eco / boutique": "Éco / boutique",
    "Eco / plein air": "Éco / plein air",
    "Eco / marche local": "Éco / marché local",
    "Eco / cosmetiques": "Éco / cosmétiques",
    "Eco / apothicaire": "Éco / apothicaire",
    "Eco / sante": "Éco / santé",
    "Eco / bien-etre": "Éco / bien-être",
    "Eco / friperie": "Éco / friperie",
    "Eco / grande boutique verte": "Éco / grande boutique verte",
    "Magasin general ecoresponsable": "Magasin général écoresponsable",
    "Magasin general eco": "Magasin général éco",
    "Epicerie zero dechet": "Épicerie zéro déchet",
    "Epicerie fine / terroir": "Épicerie fine / terroir",
    "Epicerie bio / vrac": "Épicerie bio / vrac",
    "Epicerie bio / eco": "Épicerie bio / éco",
    "Epicerie sante / eco": "Épicerie santé / éco",
    "Epicerie / vrac": "Épicerie / vrac",
    "Epicerie vrac": "Épicerie vrac",
    "Alimentation sante / bio": "Alimentation santé / bio",
    "Artisans quebecois": "Artisans québécois",
    "Artisans et designers quebecois": "Artisans et designers québécois",
    "Cadeaux / produits quebecois": "Cadeaux / produits québécois",
    "Terroir / cadeaux quebecois": "Terroir / cadeaux québécois",
    "Terroir / epicerie fine": "Terroir / épicerie fine",
    "Terroir / produits regionaux": "Terroir / produits régionaux",
    "Metiers d'art": "Métiers d'art",
    "Artisans / metiers d'art": "Artisans / métiers d'art",
    "Artisanat / materiaux creatifs": "Artisanat / matériaux créatifs",
    "Boutique de musee": "Boutique de musée",
    "Librairie independante": "Librairie indépendante",
    "Marche / eco": "Marché / éco",
    "Cafe-boutique eco": "Café-boutique éco",
    "Sante naturelle / boutique": "Santé naturelle / boutique",
    "Cooperative artisans": "Coopérative artisans",
    "Coop artisans": "Coop artisans",
    "Chasse et peche": "Chasse et pêche",
    "Deco / maison": "Déco / maison",
    "Fleuriste / jardin": "Fleuriste / jardin",
    "Boissons / brassage": "Boissons / brassage",
    "Metropole": "Métropole",
    "Region eloignee": "Région éloignée",
    "Zone deja couverte": "Zone déjà couverte",
    "Zone a ouvrir": "Zone à ouvrir",
    "Recherche web ciblee": "Recherche web ciblée",
    "A - approcher en premier": "A - approcher en premier",
    "B - releve": "B - relève",
    "C - reserve": "C - réserve",
}
def fr(v):
    return ACCENTS.get(v, v)

# --------------------------------------------------------------------- chiffrier
wb = Workbook()
TITRE = PatternFill("solid", fgColor="1F4E3D")
FTITRE = Font(bold=True, color="FFFFFF", size=11)
BORD = Border(*[Side(style="thin", color="D9D9D9")] * 4)

def feuille(ws, entetes, data, largeurs):
    ws.append(entetes)
    for i, c in enumerate(ws[1], 1):
        c.fill, c.font = TITRE, FTITRE
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for d in data:
        ws.append(d)
    for i, w in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = BORD
            c.alignment = Alignment(vertical="top", wrap_text=False)

# --- Lisez-moi
ws = wb.active
ws.title = "Lisez-moi"
lignes = [
    ["Points de vente Lasclay – liste de prospection pancanadienne", ""],
    ["", ""],
    ["Objectif du réseau", "100 à 200 points de vente actifs, dans des zones peuplées."],
    ["Modèle", "Consignation: Lasclay garde la propriété du stock jusqu'à la vente. Le détaillant n'avance pas d'argent, il engage de la tablette et de l'attention."],
    ["Ce que ça change", "Le risque d'inventaire reste chez nous. Les critères qui comptent deviennent l'achalandage, la rotation et la qualité de la reddition de comptes."],
    ["Règle d'exclusivité", "Un seul détaillant retenu par zone. Les candidats sont classés par rang: on approche le rang 1, puis le 2, puis le 3."],
    ["Règle de distance", "Le découpage en zones vise à ce que personne n'ait plus d'une heure de route pour atteindre un point de vente."],
    ["Nombre de zones", f"{len(zones)} zones couvrant les régions habitées du Canada."],
    ["Candidats listés", f"{len(retenus)}: {sum(1 for r in retenus if r['prov'] == 'QC')} au Québec et {sum(1 for r in retenus if r['prov'] != 'QC')} ailleurs au Canada."],
    ["Profondeur", "Plusieurs candidats par zone. C'est un bassin de prospection, pas une liste de détaillants approuvés."],
    ["", ""],
    ["Comment lire la feuille Candidats", ""],
    ["Priorité A", "Rang 1 dans une zone non encore couverte: à contacter en premier."],
    ["Priorité B", "Rang 2 ou 3 dans une zone non couverte: relève si le rang 1 refuse."],
    ["Priorité C", "Zone déjà couverte par un détaillant Lasclay, ou rang 4 et plus."],
    ["Pointage", "Archétype du commerce (jusqu'à 40) + population de la zone (jusqu'à 30) + coordonnées complètes (jusqu'à 25) + zone à ouvrir (15) + validation à la lecture (10)."],
    ["Priorité aux 50 000+", "Chaque zone reçoit d'abord son meilleur candidat, pour que la règle de l'heure de route tienne partout. Le reste des places va aux zones de 50 000 habitants et plus avant les petites."],
    ["Population", "Population de la ville-ancre (Wikidata). Elle sert d'ordre de grandeur du marché de la zone, pas de compte exact des habitants desservis."],
    ["Montréal", "Le centre et le Plateau sont couverts par Les Défricheuses. L'île est découpée: l'Ouest-de-l'Île et l'Est sont des zones distinctes, ouvertes. La Rive-Sud et la Rive-Nord aussi."],
    ["Rive-Sud", "La boutique du parc des Îles-de-Boucherville est saisonnière et ne porte qu'une partie de la gamme: la zone reste à ouvrir pour un détaillant à l'année."],
    ["", ""],
    ["Le tri appliqué", ""],
    ["Contact obligatoire", "Toute fiche sans téléphone, courriel, page sociale ni site web vivant a été retirée: elle n'est pas exploitable."],
    ["Chaînes", "Une enseigne présente dans quatre zones ou plus est retirée: elle achète par centrale et ne fera pas de consignation."],
    ["Archétypes écartés", "Fournitures de bricolage, friperies et maroquineries. Les friperies ne prennent pas de neuf en consignation; la maroquinerie nous concurrence sur le sac."],
    ["Archétypes conditionnels", "Alcool, sport et fleuriste ne sont gardés que si le nom porte un signal de spécialité (cave, vignoble, microbrasserie, plein air, outfitter, jardin, pépinière...)."],
    ["Signaux du nom", "Les mots artisan, local, coop, marché, terroir, ferme, atelier, vrac, éco, général, mercantile, créateur et les noms de région donnent un bonus."],
    ["Ce qui reste à faire", "Les zones marquées AUCUN candidat dans la feuille Couverture demandent une recherche manuelle: les données publiques n'y donnent aucun commerce joignable."],
    ["", ""],
    ["Sources", ""],
    ["Recherche web ciblée", "Annuaires écoresponsables, répertoires d'artisans, presse régionale, offices de tourisme régionaux, sites des boutiques. Archétype validé à la lecture."],
    ["Sites des boutiques", "3213 sites interrogés pour en extraire téléphone, courriel et adresse (données structurées puis pied de page)."],
    ["OpenStreetMap", "Moisson Overpass des commerces canadiens correspondant aux archétypes retenus, dans un rayon de 30 km autour de chaque ville-ancre."],
    ["", ""],
    ["Limites à connaître", ""],
    ["Champs vides", "Une case vide veut dire non vérifiée, pas inexistante. Ne rien inventer: valider avant tout envoi."],
    ["Colonne Ville", "Vide quand la source ne la donne pas. La zone et l'adresse situent quand même le commerce."],
    ["Fraîcheur", "Des commerces peuvent avoir fermé, déménagé ou changé de vocation. Vérifier avant d'approcher."],
    ["Bruit de la moisson", "Les étiquettes OpenStreetMap sont larges: quelques commerces listés ne conviendront pas à Lasclay. Le tri final reste humain."],
    ["Adresses extraites des sites", "Récupérées automatiquement dans les pieds de page. Fiables en général, à relire avant un envoi postal."],
]
for l in lignes: ws.append(l)
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 105
ws["A1"].font = Font(bold=True, size=14, color="1F4E3D")
for r in ws.iter_rows(min_row=1, max_col=1):
    if r[0].value and r[0].row > 2: r[0].font = Font(bold=True)
for r in ws.iter_rows(min_row=2, max_col=2):
    r[1].alignment = Alignment(wrap_text=True, vertical="top")

# --- Reseau actuel
ws = wb.create_sheet("Réseau actuel")
feuille(ws, ["Nom", "Nature du commerce", "Ville", "Prov", "Adresse", "Code postal",
             "Téléphone", "Courriel", "Site web", "Produits en boutique"],
        ACTUELS, [34, 34, 18, 6, 34, 12, 15, 26, 34, 46])

# --- Candidats
ws = wb.create_sheet("Candidats")
data = [[fr(r["priorite"]), r["rang"], r["zone"], r["ville"], r["prov"], r["nom"], fr(r["type"]),
         r["adresse"], r["cp"], r["tel"], r["courriel"], r["fb"] or r["ig"], r["web"],
         moyens_de_contact(r), r["score"], r["pop"] or "", fr(tranche_population(r["pop"])),
         fr(r["statut_zone"]), fr(r["source"])] for r in retenus]
feuille(ws, ["Priorité", "Rang zone", "Zone (1 seul détaillant)", "Ville", "Prov", "Nom",
             "Archétype", "Adresse", "Code postal", "Téléphone", "Courriel",
             "Page sociale", "Site web", "Moyens de contact", "Score",
             "Population de la zone", "Tranche de population",
             "Statut de la zone", "Source"],
        data, [22, 10, 30, 20, 6, 34, 26, 34, 12, 15, 30, 34, 36, 24, 8, 14, 20, 20, 20])

# --- Couverture
ws = wb.create_sheet("Couverture par zone")
cov = []
for z in zones:
    cands = par_zone.get(z["zone"], [])
    couverte = cle_zone(z["zone"]) in {cle_zone(x) for x in ZONES_COUVERTES}
    if couverte: statut = "Couverte - detaillant en place"
    elif cands: statut = "A ouvrir - candidats identifies"
    else: statut = "A ouvrir - AUCUN candidat trouve"
    cov.append([z["prov"], z["zone"], z["ancre"], z.get("pop", 0) or "", statut, len(cands),
                cands[0]["nom"] if cands else "", cands[0]["tel"] if cands else "",
                cands[0]["courriel"] if cands else ""])
cov.sort(key=lambda r: (r[0], r[1]))
feuille(ws, ["Prov", "Zone", "Ville-ancre", "Population", "Statut", "Candidats trouvés",
             "Meilleur candidat", "Téléphone", "Courriel"],
        cov, [6, 34, 22, 13, 32, 16, 34, 15, 30])

# --- Grille de qualification
ws = wb.create_sheet("Grille de qualification")
grille = [
    ["1. Récit", "La boutique sait-elle raconter d'où vient un produit?", "Éliminatoire",
     "L'asclépiade ne se vend pas toute seule. Sans quelqu'un qui explique le monarque et la fibre, le produit dort sur la tablette."],
    ["2. Indépendance", "Commerce indépendant, pas une succursale de chaîne", "Éliminatoire",
     "Une chaîne achète par centrale et ne fait pas de consignation."],
    ["3. Clientèle", "Clientèle nature, plein air, jardin, achat local ou cadeau", "Éliminatoire",
     "Correspond aux clientèles Odette, Suzanne et Gilbert, Alexandra."],
    ["4. Achalandage", "Passage suffisant, en zone habitée ou touristique", "Fort",
     "En consignation, c'est nous qui portons l'inventaire. Un commerce tranquille immobilise notre stock sans le vendre."],
    ["5. Vitrine", "Espace pour une présentation dédiée avec une affichette explicative", "Fort",
     "Ce que le détaillant met dans l'entente, ce n'est pas de l'argent, c'est de la tablette et de l'attention. C'est ça qu'il faut négocier."],
    ["6. Rotation", "Quel volume la boutique prévoit-elle écouler par saison?", "Fort",
     "Le vrai coût de la consignation est le stock immobilisé. Mieux vaut trois boutiques qui roulent que dix qui gardent une boîte au sous-sol."],
    ["7. Saison", "Ouvert à l'année, ou complémentaire à notre saisonnalité", "Fort",
     "Une boutique estivale porte les glacières et les sacs à lunch; une boutique d'hiver, les mitaines."],
    ["8. Exclusivité", "Accepte d'être le seul détaillant de sa zone", "Fort",
     "C'est la contrepartie offerte au commerce, et ce qui justifie qu'il s'investisse."],
    ["9. Reddition", "Capable de faire un rapport de ventes périodique et un inventaire de fin de saison", "Fort",
     "En consignation, sans rapport fiable on ne sait ni ce qui se vend, ni ce qu'on possède encore. C'est le point de friction habituel du modèle."],
    ["10. Retours", "Accepte les modalités de réassort et de retour des invendus", "Moyen",
     "À clarifier par écrit dès le départ: qui paie le transport de retour, dans quel état, à quelle date."],
    ["11. Numérique", "Présence en ligne active (site, réseaux sociaux, infolettre)", "Moyen",
     "Un détaillant qui publie prolonge notre portée sans coût d'acquisition."],
    ["12. Prix", "Accepte le prix de détail suggéré", "Moyen",
     "Évite qu'un même produit se retrouve à deux prix trop différents de lasclay.com."],
    ["13. Produits d'entrée", "Commence par les produits sans taille ni saison", "Méthode",
     "Sacs à lunch, manchons, glacières: ils tournent à l'année et n'obligent pas la boutique à gérer un tableau de tailles. C'est ce que le réseau actuel prend spontanément."],
    ["14. Semences", "Prend les semences d'asclépiade", "Méthode",
     "Petit prix, achat d'impulsion, et c'est le coeur du récit du monarque. Ça ouvre la conversation en magasin."],
]
feuille(ws, ["Critère", "Question à poser", "Poids", "Pourquoi ça compte"],
        grille, [24, 58, 16, 78])

out = os.path.join(BASE, "points-de-vente-lasclay.xlsx")
wb.save(out)

n_zones_ok = sum(1 for z in zones if par_zone.get(z["zone"]))
print(f"Candidats retenus : {len(retenus)} (bassin total {len(rows)})")
print(f"Zones avec au moins un candidat : {n_zones_ok}/{len(zones)}")
print(f"Avec telephone : {sum(1 for r in retenus if r['tel'])}")
print(f"Avec courriel  : {sum(1 for r in retenus if r['courriel'])}")
print(f"Avec adresse   : {sum(1 for r in retenus if r['adresse'])}")
print(f"Fichier : {out}")

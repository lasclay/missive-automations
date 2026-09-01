#!/usr/bin/env python3
"""Ecrit le navigateur de brouillons : une page sans JavaScript.

L'apercu de fichier des applications mobiles n'execute pas les scripts. Une
page qui construit sa liste au chargement s'y affiche donc vide sous l'entete,
ce qui est exactement ce qu'on a vu. Les 250 fiches sont maintenant ecrites en
dur, et les filtres sont des boutons radio caches que le CSS lit : tout
fonctionne script desactive, hors ligne, dans n'importe quel navigateur.

Deux sorties, meme contenu :
  · brouillons-medias.html         fichier autonome, avec son entete
  · <scratchpad>/brouillons-medias.html   version publiee comme artefact
"""
import html as H
import os
import re
import sys

import openpyxl

VERSION = 12
SOURCE = "Lasclay_v2.xlsx"
AUTONOME = "brouillons-medias.html"
ARTEFACT = os.path.join(
    "/tmp/claude-0/-home-user-missive-automations",
    "dcfcba14-3889-5dfd-96e3-63264ff80ae0/scratchpad/brouillons-medias.html")

ANGLES = {
    "A": "a déjà couvert Lasclay", "B": "média régional", "C": "agriculture et filière",
    "D": "environnement et monarques", "E": "affaires et manufacturier",
    "F": "plein air et équipement", "G": "style de vie et design",
    "H": "presse anglophone", "I": "radio et télévision",
    "J": "a couvert l'asclépiade, jamais Lasclay", "—": "doublon, ne pas envoyer",
}
# Les brouillons ecrits un par un, par opposition a ceux qui sortent du gabarit.
MAIN = re.compile(r"Mékinac|Saint-Tite|Téléjournal|Granby|soyer du Québec")

CSS = """
:root{--paper:#F1F3EE;--surface:#FBFCF9;--ink:#161C17;--muted:#5C6659;--faint:#7C8578;
--rule:#D6DCD0;--rule-soft:#E5E9E0;--accent:#3F5B45;--accent-soft:#E3EADF;--flag:#A93D18;--flag-soft:#F6E6DE;
--sans:"Public Sans",-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
--serif:"Newsreader",Georgia,serif;--mono:"JetBrains Mono",ui-monospace,Menlo,monospace}
@media (prefers-color-scheme:dark){:root:not([data-theme="light"]){--paper:#12160F;--surface:#1A2017;
--ink:#E7EBE0;--muted:#9AA394;--faint:#7F887A;--rule:#2C342A;--rule-soft:#232A21;--accent:#9DBE9F;
--accent-soft:#222C22;--flag:#E28558;--flag-soft:#2E2018}}
:root[data-theme="dark"]{--paper:#12160F;--surface:#1A2017;--ink:#E7EBE0;--muted:#9AA394;--faint:#7F887A;
--rule:#2C342A;--rule-soft:#232A21;--accent:#9DBE9F;--accent-soft:#222C22;--flag:#E28558;--flag-soft:#2E2018}
*{box-sizing:border-box}
body{background:var(--paper);color:var(--ink);font-family:var(--sans);font-size:15px;
line-height:1.6;-webkit-text-size-adjust:100%}
h1{font-family:var(--serif);font-weight:600;margin:0;font-size:22px;letter-spacing:-.015em;line-height:1.15}
p{margin:0}
.top{border-bottom:1px solid var(--rule);background:var(--surface)}
.top-in{max-width:1280px;margin:0 auto;padding:14px 24px;display:flex;flex-direction:column;gap:12px}
.ttl{display:flex;align-items:baseline;gap:16px;flex-wrap:wrap}
.sub{font-family:var(--mono);font-size:12px;color:var(--muted);font-variant-numeric:tabular-nums}
.sub b{color:var(--flag)}
.bar{display:flex;gap:8px;flex-wrap:wrap;align-items:center}
.chip{font-size:12.5px;font-weight:500;padding:5px 11px;border:1px solid var(--rule);
background:var(--paper);color:var(--muted);border-radius:99px;cursor:pointer;font-family:var(--mono);
display:inline-block;user-select:none}
.sep{width:1px;height:20px;background:var(--rule);margin:0 3px}
main{max-width:1280px;margin:0 auto;padding:20px 24px 90px;display:flex;flex-direction:column;gap:10px}
.note{background:var(--flag-soft);border-left:3px solid var(--flag);border-radius:3px;padding:13px 16px;
font-size:13.5px;color:var(--muted);line-height:1.55}
.note b{color:var(--ink);font-weight:500}
details{background:var(--surface);border:1px solid var(--rule);border-radius:3px;overflow:hidden}
details[open]{border-color:var(--accent)}
summary{padding:12px 16px;cursor:pointer;display:grid;
grid-template-columns:26px minmax(150px,1.1fr) minmax(120px,1fr) minmax(160px,1.3fr) auto;
gap:14px;align-items:center;list-style:none}
summary::-webkit-details-marker{display:none}
summary:hover{background:var(--accent-soft)}
.ang{font-family:var(--mono);font-size:12px;font-weight:500;width:24px;height:24px;display:grid;
place-items:center;background:var(--accent-soft);color:var(--accent);border-radius:2px}
.ang.no{background:var(--flag-soft);color:var(--flag)}
.nm{font-weight:500;font-size:14.5px}
.md{font-size:13px;color:var(--muted)}
.em{font-family:var(--mono);font-size:11.5px;color:var(--faint);word-break:break-all}
.em.none{color:var(--flag)}
.tags{display:flex;gap:5px;justify-self:end;flex-wrap:wrap}
.tg{font-family:var(--mono);font-size:10.5px;padding:2px 7px;border-radius:2px;
background:var(--rule-soft);color:var(--faint);white-space:nowrap}
.tg.hot{background:var(--flag-soft);color:var(--flag)}
.tg.hand{background:var(--accent-soft);color:var(--accent)}
.body{padding:0 16px 16px;display:flex;flex-direction:column;gap:11px;
border-top:1px solid var(--rule-soft);margin-top:2px;padding-top:14px}
.meta{display:flex;gap:16px;flex-wrap:wrap;font-size:12.5px;color:var(--faint)}
.meta span b{color:var(--muted);font-weight:400}
.subj{font-family:var(--mono);font-size:12.5px;color:var(--muted)}
.subj b{color:var(--ink);font-weight:400}
.warn{background:var(--flag-soft);color:var(--flag);border-radius:3px;padding:9px 13px;font-size:13px}
pre{margin:0;background:var(--paper);border:1px solid var(--rule);border-radius:3px;padding:16px 18px;
font-family:var(--mono);font-size:12.8px;line-height:1.75;color:var(--ink);white-space:pre-wrap;overflow-x:auto}
/* Les filtres sont des boutons radio caches : le CSS masque ce qui ne
   correspond pas. Aucun script, donc rien a executer pour que ca marche. */
.f{position:absolute;width:1px;height:1px;opacity:0;pointer-events:none}
.f:checked + .top .chip[for]{opacity:1}
FILTRES
@media(max-width:700px){
.top-in{padding:12px 14px;gap:10px}h1{font-size:19px}.sub{font-size:11.5px}
main{padding:14px 14px 60px}.chip{font-size:12px;padding:4px 9px}.sep{display:none}
summary{padding:11px 13px;grid-template-columns:26px 1fr;gap:8px}
.md,.em,.tags{grid-column:2}.tags{justify-self:start}
.body{padding:0 13px 14px;padding-top:12px}
pre{font-size:12px;padding:12px 13px;line-height:1.7}
.note{font-size:13px;padding:12px 13px}}
"""


def index(ws):
    return {c.value: i for i, c in enumerate(ws[1])}


def fiches():
    wb = openpyxl.load_workbook(SOURCE)
    ws = wb["Liste chaude (34)"]
    h = index(ws)
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[h["Nom"]]:
            continue
        out.append({"l": "chaude", "n": r[h["Nom"]], "m": r[h["Média"]] or "",
                    "e": r[h["Courriel"]] or "", "d": str(r[h["Dernier contact"]] or ""),
                    "s": r[h["Sujet du dernier échange"]] or "", "a": r[h["Angle"]] or "",
                    "o": r[h["Objet suggéré"]] or "", "r": r[h["Registre"]] or "",
                    "p": r[h["Notes"]] or "", "t": r[h["Brouillon"]] or ""})

    ws = wb["Liste froide FPJQ (217)"]
    h = index(ws)
    for r in ws.iter_rows(min_row=2, values_only=True):
        nom = " ".join(x for x in (r[h["Prénom"]], r[h["Nom"]]) if x)
        if not nom:
            continue
        out.append({"l": "froide", "n": nom, "m": r[h["Média"]] or "",
                    "e": r[h["Courriel"]] or "", "d": r[h["Priorité"]] or "",
                    "s": r[h["Secteurs pertinents"]] or "", "a": r[h["Angle"]] or "",
                    "o": r[h["Objet suggéré"]] or "", "r": r[h["Région"]] or "",
                    "p": r[h["Précaution"]] or "", "t": r[h["Brouillon"]] or ""})

    for c in out:
        c["dead"] = c["t"].startswith("NE PAS ENVOYER")
        c["hand"] = (c["l"] == "chaude" or c["a"] == "J" or bool(MAIN.search(c["t"])))
    return out


def e(s):
    return H.escape(str(s or ""), quote=False)


def fiche(c):
    classes = ["l-" + c["l"], "a-" + (c["a"] or "x")]
    if c["hand"]:
        classes.append("main")
    if not c["e"]:
        classes.append("noem")
    tags = ('<span class="tg hot">chaude</span>' if c["l"] == "chaude"
            else f'<span class="tg">{e(c["d"])}</span>')
    if c["hand"] and not c["dead"]:
        tags += '<span class="tg hand">à la main</span>'
    meta = [f'<span><b>Angle {e(c["a"])}</b> · {e(ANGLES.get(c["a"], ""))}</span>']
    if c["r"]:
        meta.append(f'<span><b>{"Registre" if c["l"] == "chaude" else "Région"}</b> {e(c["r"])}</span>')
    if c["s"]:
        meta.append(f'<span><b>{"Dernier échange" if c["l"] == "chaude" else "Secteurs"}</b> {e(c["s"])}</span>')
    corps = [f'<div class="meta">{"".join(meta)}</div>']
    if c["o"]:
        corps.append(f'<div class="subj">Objet : <b>{e(c["o"])}</b></div>')
    if c["p"]:
        corps.append(f'<div class="warn">{e(c["p"])}</div>')
    corps.append(f'<pre>{e(c["t"])}</pre>')
    return (f'<details class="{" ".join(classes)}"><summary>'
            f'<span class="ang{" no" if c["dead"] else ""}">{e(c["a"]) or "?"}</span>'
            f'<span class="nm">{e(c["n"])}</span>'
            f'<span class="md">{e(c["m"])}</span>'
            f'<span class="em{"" if c["e"] else " none"}">{e(c["e"]) or "adresse à trouver"}</span>'
            f'<span class="tags">{tags}</span></summary>'
            f'<div class="body">{"".join(corps)}</div></details>')


def page(contacts):
    angles = sorted({c["a"] for c in contacts if c["a"] and c["a"] != "—"})
    filtres = [("tous", "tous", len(contacts), None)]
    filtres += [("chaude", "liste chaude", sum(1 for c in contacts if c["l"] == "chaude"), "l-chaude"),
                ("froide", "liste froide", sum(1 for c in contacts if c["l"] == "froide"), "l-froide"),
                ("main", "écrits à la main", sum(1 for c in contacts if c["hand"]), "main"),
                ("noem", "sans courriel", sum(1 for c in contacts if not c["e"]), "noem")]
    for a in angles:
        filtres.append((f"a-{a}", a, sum(1 for c in contacts if c["a"] == a), f"a-{a}"))
    # Un filtre qui ne ramene rien n'a pas a occuper une place dans la barre.
    filtres = [f for f in filtres if f[2]]

    regles = "\n".join(f"#f-{ident}:checked ~ main details:not(.{cls}),"
                       for ident, _, _, cls in filtres if cls)
    regles = regles.rstrip(",") + "{display:none}\n" + "\n".join(
        f'#f-{ident}:checked ~ .top label[for="f-{ident}"]'
        "{background:var(--accent-soft);border-color:var(--accent);color:var(--accent)}"
        for ident, _, _, _ in filtres)

    radios = "\n".join(
        f'<input class="f" type="radio" name="filtre" id="f-{ident}"'
        f'{" checked" if ident == "tous" else ""}>' for ident, _, _, _ in filtres)
    puces = []
    for i, (ident, nom, n, _) in enumerate(filtres):
        if ident == "a-" + angles[0]:
            puces.append('<span class="sep"></span>')
        puces.append(f'<label class="chip" for="f-{ident}">{e(nom)} · {n}</label>')

    note = (
        '<div class="note"><b>Version 12.</b> Le lien du média kit est dans les 247 '
        "brouillons, juste avant le paragraphe des bénéfices : un journaliste qui envisage "
        "un sujet veut savoir tout de suite s'il aura des images. Le corps suit le squelette "
        "que Gabriel a corrigé lui-même, et Larocque et Pouliot sont repris mot pour mot."
        "<br><br><b>Rien ne part sans relecture.</b> Le proxy Missive dépose des brouillons : "
        "un appel par journaliste, en <code>to</code> seul, depuis media@lasclay.com, suivi "
        "des ouvertures et des clics désactivé, environ 60 par jour."
        "<br><br><b>Pour chercher un nom</b>, utilise la recherche du navigateur : sur "
        "iPhone, le menu Partager puis « Rechercher dans la page ».</div>")

    return (
        "<title>Brouillons médias Lasclay</title>\n"
        '<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>\n'
        '<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Newsreader:'
        "opsz,wght@6..72,400;6..72,600&family=Public+Sans:wght@400;500;700&family="
        'JetBrains+Mono:wght@400;500&display=swap">\n'
        f"<style>{CSS.replace('FILTRES', regles)}</style>\n\n"
        f"{radios}\n"
        '<div class="top"><div class="top-in">\n'
        f'  <div class="ttl"><h1>Brouillons médias</h1>'
        f'<div class="sub">{len(contacts)} contacts · v{VERSION} · diffusion '
        "<b>jeu. 17 sept., 20 h</b></div></div>\n"
        f'  <div class="bar">{"".join(puces)}</div>\n'
        "</div></div>\n\n<main>\n" + note + "\n"
        + "\n".join(fiche(c) for c in contacts) + "\n</main>\n")


def main():
    contacts = fiches()
    corps = page(contacts)
    open(ARTEFACT, "w", encoding="utf-8").write(corps)
    open(AUTONOME, "w", encoding="utf-8").write(
        '<!doctype html>\n<html lang="fr">\n<head>\n<meta charset="utf-8">\n'
        '<meta name="viewport" content="width=device-width,initial-scale=1">\n'
        "<style>body{margin:0}</style>\n"
        + corps.replace("</style>\n\n", "</style>\n</head>\n<body>\n", 1)
        + "</body>\n</html>\n")
    kit = sum(1 for c in contacts if "1pyCUbfHYQhpXXl4FoCC2RCFXKRvGS5Zr" in c["t"])
    print(f"{len(contacts)} fiches, {kit} avec le média kit, aucun script")


if __name__ == "__main__":
    sys.exit(main())

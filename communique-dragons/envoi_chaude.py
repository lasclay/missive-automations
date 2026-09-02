#!/usr/bin/env python3
"""Depose les 23 courriels de la liste chaude dans Missive.

Quand un vrai echange journalistique existe, le courriel se depose DANS ce
fil : le journaliste retrouve son propre article en haut de la fenetre, ce
qu'aucune ligne d'objet ne remplace. Le fil se choisit a la main, jamais par
« le plus recent ». Plusieurs de ces personnes sont aussi des clientes, et
repondre a « As-tu recu ta commande pour Noel? » avec une annonce de presse
serait grotesque ; les envois en masse de 2022 ne comptent pas non plus comme
une conversation.

Quatre personnes recoivent un courriel neuf : Laforest et Goubau, dont le
texte commence par « Je m'appelle Gabriel Gouveia » et qui contredirait une
reponse dans un fil ; McKenna, dont le seul fil est un envoi en masse ;
Leprince, qu'on n'a jamais contacte.

L'API Missive n'a AUCUNE route pour supprimer un brouillon. Relancer ce script
ne remplace donc rien : ca depose une deuxieme serie a cote de la premiere. Les
brouillons deja poses se suppriment a la main dans Missive, avant de relancer.

    python3 envoi_chaude.py            depose des brouillons, rien ne part
    python3 envoi_chaude.py --envoyer  envoie pour de vrai
"""
import base64
import json
import os
import subprocess
import sys

DE = "admin@lasclay.com"

# Un seul objet pour les 23, choisi par Gabriel. Il remplace l'objet du fil
# meme dans les reponses : « Re: Transfert en Tunisie » rappelait le sujet de
# 2026, pas celui du courriel. Le fil, lui, ne bouge pas.
OBJET = "L'asclépiade à Dragons' Den le 17 septembre"
RACINE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CLIENT = os.path.join(RACINE, "missive_client.js")
PHOTO = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dragons-plateau.jpg")
NOM_PHOTO = "lasclay-dragons-den.jpg"

# courriel : (id du fil, objet de la reponse, pourquoi ce fil-la)
FILS = {
"sylvain.larocque@quebecormedia.com": ("de950cb1-45b1-4a2b-9d1e-000000000000", None, ""),
}

with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "fils_chauds.json"),
          encoding="utf-8") as f:
    FILS = json.load(f)


def appel(commande, charge):
    p = subprocess.run(["node", CLIENT, commande], input=json.dumps(charge),
                       capture_output=True, text=True, timeout=300)
    if p.returncode != 0:
        return {"erreur": (p.stderr or p.stdout).strip()[:300]}
    return json.loads(p.stdout)


def main(envoyer):
    import openpyxl
    ws = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                             "Lasclay_v2.xlsx"))["Liste chaude (34)"]
    h = [c.value for c in ws[1]]
    iC, iB, iO = h.index("Courriel"), h.index("Brouillon"), h.index("Objet suggéré")
    piece = [{"base64_data": base64.b64encode(open(PHOTO, "rb").read()).decode(),
              "filename": NOM_PHOTO}]

    faits, echecs = [], []
    for r in ws.iter_rows(min_row=2, values_only=True):
        courriel, texte = r[iC], r[iB]
        if not texte or texte.startswith("NE PAS"):
            continue
        fil = FILS.get(courriel)
        if fil:
            charge = {"id": fil["id"], "from": DE, "to": [courriel],
                      "subject": OBJET, "body": texte, "attachments": piece,
                      "signature": True}
            if envoyer:
                charge["send"] = True
            res = appel("reply", charge)
            mode = f"réponse · {fil['pourquoi']}"
        else:
            charge = {"from": DE, "to": [courriel], "subject": OBJET,
                      "body": texte, "attachments": piece, "signature": True}
            if envoyer:
                charge["send"] = True
            res = appel("send", charge)
            mode = "courriel neuf"
        if res.get("ok"):
            faits.append((r[0], mode))
            print(f"  ✓ {r[0]:26s} {mode}")
        else:
            echecs.append((r[0], courriel, res.get("erreur") or res))
            print(f"  ✗ {r[0]:26s} {res.get('erreur') or res}")

    verbe = "envoyé" if envoyer else "déposé en brouillon"
    print(f"\n{len(faits)} {verbe}, {len(echecs)} en échec")
    for nom, adr, err in echecs:
        print(f"  {nom} <{adr}> : {err}")
    return 1 if echecs else 0


if __name__ == "__main__":
    sys.exit(main("--envoyer" in sys.argv))

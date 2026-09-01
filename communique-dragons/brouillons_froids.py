#!/usr/bin/env python3
"""Les 219 brouillons de la liste froide, dans le registre de Gabriel.

Meme squelette que la liste chaude, sauf que l'ouverture ne peut pas rappeler un
article : elle part de ce qui concerne la personne, region ou sujet.

Deux differences assumees avec la liste chaude :

- Salutation « Bonjour Prenom Nom ». Deviner M. ou Mme sur 219 personnes qu'on
  ne connait pas produirait des erreurs; le prenom et le nom evitent la question.
- Les accroches regionales sont tournees vers ce qui existe aujourd'hui, pas vers
  la faillite de 2018. Le fait reste vrai, le ton reste celui d'une bonne
  nouvelle.

    python3 brouillons_froids.py <chiffrier.xlsx> <sortie.xlsx> <sortie.md>
"""

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from voix_gabriel import VIDEO, ANNONCE, BENEFICE, OBJETS, assembler

QUI = ("Je m'appelle Gabriel Gouveia, fondateur de Lasclay. On isole des vêtements d'hiver "
       "avec une mauvaise herbe, l'asclépiade, qu'on cultive pour sauvegarder un pollinisateur "
       "emblématique et menacé : le papillon monarque.")

# L'ouverture porte tout le pitch d'un coup : le produit, la plante, la mission.
# Elle rend le paragraphe des deux missions redondant dans la liste froide, donc
# il saute : le courriel raccourcit d'autant. Les « petits cochons » designent
# les gousses, pas la soie : la phrase doit le dire dans cet ordre.
CONTEXTE = ("Ses gousses, les « petits cochons » que les agriculteurs arrachent de leurs champs "
            "depuis 50 ans, sont remplies d'une soie creuse, très légère et naturellement "
            "hydrophobe. On la transforme en isolant à Québec, et on en fait des manteaux, des "
            "mitaines, des tuques et des sacs isothermes vendus au Canada et aux États-Unis.")

MISSIONS_2 = ("Mes 2 grandes missions : faire connaître l'asclépiade et sauvegarder les "
              "monarques par la culture de l'asclépiade.")

PAS = "Le 17 septembre prochain, on va faire un énorme pas dans la bonne direction :"

RARETE = ("Aller présenter notre entreprise et sa mission à la télévision nationale est une "
          "opportunité qui arrive bien rarement. Je voulais vous en faire part et qui sait, "
          "peut-être vous inspirer un sujet.")

CLOTURE = "Au plaisir et n'hésitez pas à me contacter si vous avez des questions."

# --- pourquoi je vous ecris a vous ------------------------------------------
# Une seule raison par courriel, et elle doit etre vraie. La region ne gagne que
# la ou l'asclepiade a une histoire locale reelle : « notre atelier est dans
# Limoilou » n'est pas une accroche locale pour un journal de Charlevoix.

REGIONS = {
    "Mauricie":
        "Je vous écris entre autres parce que l'asclépiade a eu son grand moment industriel "
        "chez vous : l'usine de Saint-Tite achetait 90 % des récoltes du Québec avant que la "
        "filière se casse en 2018. Des producteurs de la Mauricie cultivent encore, et on "
        "continue d'acheter leur récolte.",
    "Centre-du-Québec":
        "Je vous écris entre autres parce que l'asclépiade se cultive dans votre région depuis "
        "la première vague de 2013, et que plusieurs des producteurs qui ont tenu bon nous "
        "vendent encore leur récolte.",
    "Estrie":
        "Je vous écris entre autres parce qu'une des premières usines de transformation de la "
        "fibre était à Granby, et que l'Estrie compte encore des producteurs d'asclépiade.",
    "Montérégie":
        "Je vous écris entre autres parce que la Montérégie est une des régions où l'asclépiade "
        "se cultive encore, et que c'est de champs comme ceux-là que vient la fibre qu'on "
        "transforme.",
    "Saguenay - Lac-Saint-Jean":
        "Je vous écris entre autres parce qu'un de nos fournisseurs d'asclépiade cultive au "
        "Lac-Saint-Jean depuis nos tout débuts, et qu'il l'est encore aujourd'hui. La fibre de "
        "chez vous se retrouve dans nos produits.",
}

# Sinon, la raison vient du sujet que la personne couvre.
THEMES = {
    "C": "Je vous écris parce que vous couvrez l'agriculture. La question qui compte pour les "
         "producteurs n'a pas changé depuis l'effondrement de la filière en 2018 : est-ce "
         "qu'il y a un acheteur stable au bout du champ. On achète encore la récolte de "
         "producteurs d'ici et on la transforme nous-mêmes.",
    "D": "Je vous écris parce que vous couvrez l'environnement. Notre pari est économique avant "
         "d'être militant : si l'asclépiade devient payante, les agriculteurs la gardent dans "
         "leurs champs, et les monarques retrouvent de l'habitat de reproduction.",
    "E": "Je vous écris parce que vous couvrez l'économie. On a bâti nos propres procédés de "
         "transformation parce qu'aucun sous-traitant ne voulait toucher à l'asclépiade, puis "
         "on a changé de modèle manufacturier l'an dernier pour rendre un manteau accessible à "
         "300 $. J'en ai fait une vidéo ici : " + VIDEO + ".",
    "F": "Je vous écris parce que vous couvrez le plein air. À poids égal, la littérature donne "
         "la soie d'asclépiade pour environ 10 % plus isolante que le duvet. C'est un repère de "
         "laboratoire, pas une promesse de manteau, et c'est exactement pour ça que je préfère "
         "qu'on la teste.",
    "G": "Je vous écris parce que vous couvrez l'art de vivre et la consommation. Le contraste "
         "se photographie bien : la gousse dans le champ, la soie blanche dans la main, le "
         "manteau porté en ville.",
    "I": "Je vous écris parce que le sujet se raconte bien en ondes : une entreprise de Québec "
         "qui va expliquer à un auditoire pancanadien pourquoi la mauvaise herbe des champs de "
         "maïs peut isoler un manteau.",
    "B": "Je vous écris parce que c'est une nouvelle de chez nous qui passe au national.",
    "H": "",
}

# --- ce qu'on offre, par angle ----------------------------------------------

OFFRES = {
    "B": "Je suis disponible pour une entrevue avant ou après la diffusion.",
    "C": "Je peux parler des volumes, de la fenêtre de récolte et de ce qui reste à régler "
         "côté mécanisation, sans rien promettre que je ne peux pas tenir.",
    "D": "Je peux aussi parler de notre campagne de plantation, qui en est à sa 5e édition et "
         "qui a distribué environ 10 millions de graines en Amérique du Nord.",
    "E": "Je peux parler de ce que ça demande d'industrialiser une matière qui n'a aucune "
         "chaîne d'approvisionnement.",
    "F": "Et si vous voulez tester plutôt que me croire sur parole, je vous envoie un produit "
         "avec plaisir.",
    "G": "L'atelier de Limoilou est ouvert si vous voulez voir la matière, et j'ai du matériel "
         "photo.",
    "I": "Je suis disponible pour une entrevue la semaine du 14 septembre et le vendredi 18 au "
         "matin.",
}

# --- textes ecrits a la main ------------------------------------------------

MAIN = {
"jean-michel_leprince@radio-canada.ca": """Bonjour Jean-Michel Leprince,

Je m'appelle Gabriel Gouveia, fondateur de Lasclay. On isole des vêtements d'hiver avec une
mauvaise herbe, l'asclépiade, qu'on cultive pour sauvegarder un pollinisateur emblématique et
menacé : le papillon monarque.

Vous couvrez l'asclépiade depuis au moins 2014 : le Téléjournal, la soie d'Amérique partie
sur l'Everest, le pouvoir absorbant de la fibre sur les hydrocarbures, le lien avec le
monarque. Vous avez suivi cette histoire plus longtemps que la plupart des entreprises qui
s'y sont essayées.

On a démarré après la chute de la première filière. Six ans plus tard, on achète encore de l'asclépiade québécoise et on transforme l'isolant nous-mêmes à
Limoilou.

Le 17 septembre prochain, on va faire un énorme pas dans la bonne direction :

{ANNONCE}

Vous êtes probablement la personne au Québec qui a le plus longtemps suivi cette plante-là.
Je voulais vous en faire part et qui sait, peut-être vous inspirer un sujet : ce serait la
première fois qu'on raconte ce que la filière est devenue après la chute.

{BENEFICE}

{CLOTURE}""",

"redaction.mekinac@lebulletindeschenaux.com": """Bonjour Réjean Martin,

Je m'appelle Gabriel Gouveia, fondateur de Lasclay. On isole des vêtements d'hiver avec une
mauvaise herbe, l'asclépiade, qu'on cultive pour sauvegarder un pollinisateur emblématique et
menacé : le papillon monarque.

Vous couvrez Mékinac, donc Saint-Tite, donc l'endroit où l'asclépiade a eu son grand moment
industriel au Québec. L'usine achetait 90 % des récoltes de la province avant que la filière
se casse en 2018.

On a démarré après. Six ans plus tard, on achète encore de l'asclépiade québécoise et on la
transforme nous-mêmes à Québec.

Le 17 septembre prochain, on va faire un énorme pas dans la bonne direction :

{ANNONCE}

Je voulais vous en faire part parce que vos lecteurs ont vu la promesse de l'asclépiade de
plus près que n'importe qui. Qui sait, peut-être vous inspirer un sujet sur ce que la plante
est devenue depuis. Je peux parler des volumes, de la récolte, et de ce qui reste à régler.

{BENEFICE}

{CLOTURE}""",

"gabriel.delisle@lenouvelliste.qc.ca": """Bonjour Gabriel Delisle,

Je m'appelle Gabriel Gouveia, fondateur de Lasclay. On isole des vêtements d'hiver avec une
mauvaise herbe, l'asclépiade, qu'on cultive pour sauvegarder un pollinisateur emblématique et
menacé : le papillon monarque.

L'asclépiade a été une histoire mauricienne avant d'être la nôtre : l'usine de Saint-Tite
achetait 90 % des récoltes du Québec. On a démarré après la chute de cette filière, et six ans
plus tard on achète encore de l'asclépiade québécoise et on la
transforme nous-mêmes à Québec.

Le 17 septembre prochain, on va faire un énorme pas dans la bonne direction :

{ANNONCE}

Je voulais vous en faire part et qui sait, peut-être vous inspirer un sujet : huit ans après
Saint-Tite, il reste quelque chose de cette promesse-là, et ça se raconte.

{BENEFICE}

{CLOTURE}""",

"marie.allard@laesdebrouillards.com": """Bonjour Marie Allard,

Je m'appelle Gabriel Gouveia, fondateur de Lasclay. On isole des vêtements d'hiver avec une
mauvaise herbe, l'asclépiade, qu'on cultive pour sauvegarder un pollinisateur emblématique et
menacé : le papillon monarque.

Il y a une belle histoire de sciences dans l'asclépiade. La soie attachée à ses graines n'est
pas un poil : c'est un tube creux enduit d'une cire hydrophobe, et les fibres portent une
charge qui les fait se repousser. C'est ce qui forme le parachute autour de la graine, et
c'est ce qui emprisonne l'air. Chaque follicule en produit plus de 200.

On en fait de l'isolant pour des manteaux et des mitaines. Et c'est la seule plante que les
chenilles du monarque peuvent manger, ce qui est toute la raison d'être de l'entreprise.

Le 17 septembre prochain, on va faire un énorme pas dans la bonne direction :

{ANNONCE}

Je voulais vous en faire part et qui sait, peut-être vous inspirer un sujet. L'atelier de
Limoilou est ouvert si vos lecteurs aimeraient voir comment une gousse devient un manteau.

{BENEFICE}

{CLOTURE}""",

"__stab": """Bonjour Antoine Stab,

Je m'appelle Gabriel Gouveia, fondateur de Lasclay. On isole des vêtements d'hiver avec une
mauvaise herbe, l'asclépiade, qu'on cultive pour sauvegarder un pollinisateur emblématique et
menacé : le papillon monarque.

Vous aviez écrit sur le soyer du Québec dans Espaces en février 2015, à l'époque où la fibre
devait remplacer le duvet.

On a démarré après la chute de cette première filière. Six ans plus tard, on achète encore de l'asclépiade québécoise, on transforme l'isolant nous-mêmes, et il
existe maintenant un manteau à inserts d'asclépiade amovibles autour de 300 $.

Le 17 septembre prochain, on va faire un énorme pas dans la bonne direction :

{ANNONCE}

Je voulais vous en faire part et qui sait, peut-être vous inspirer un sujet. Et si vous
voulez tester le manteau plutôt que me croire sur parole, je vous en envoie un avec plaisir.

{BENEFICE}

{CLOTURE}""",
}

DOUBLON = ("NE PAS ENVOYER — doublon de la liste chaude. Le message écrit à la main a "
           "préséance.")


def salutation_douteuse(prenom, nom):
    """Le repertoire FPJQ contient des fiches mal formees : nom en majuscules,
    prenom et nom intervertis, plusieurs prenoms colles. On ne corrige pas en
    silence, on signale."""
    plein = f"{prenom} {nom}".strip()
    return len(plein) > 28 or (nom and nom.isupper() and len(nom) > 2)


def monter(prenom, nom, angle, region):
    """On se presente toujours : ces gens ne nous connaissent pas. L'accroche
    regionale ou thematique vient ensuite, jamais avant la presentation."""
    salut = f"Bonjour {prenom} {nom},".replace("  ", " ")
    pourquoi = REGIONS.get(region) or THEMES.get(angle) or THEMES["B"]
    offre = OFFRES.get(angle, "")
    return assembler([salut, QUI, CONTEXTE, pourquoi, PAS, ANNONCE,
                      RARETE + (" " + offre if offre else ""), BENEFICE, CLOTURE])


def main(src, dst, md):
    wb = openpyxl.load_workbook(src)
    ws = next(w for w in wb.worksheets if w.title.startswith("Liste froide"))
    ent = [c.value for c in ws[1]]
    cA = ent.index("Angle") + 1
    cO = ent.index("Objet suggéré") + 1
    cB = ent.index("Brouillon") + 1 if "Brouillon" in ent else ws.max_column + 1
    c = ws.cell(1, cB, "Brouillon")
    c.font = Font(bold=True, size=10); c.fill = PatternFill("solid", fgColor="E3EADF")
    ws.column_dimensions[ws.cell(1, cB).column_letter].width = 78

    par_angle, mains, douteux = {}, 0, []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not r[2]:
            continue
        prio, prenom, nom = r[0], (r[1] or "").strip(), (r[2] or "").strip()
        media, courriel, region, angle = r[4], (r[5] or "").strip().lower(), r[7], r[cA - 1]
        if angle in (None, "—"):
            ws.cell(i, cB, DOUBLON).alignment = Alignment(wrap_text=True, vertical="top")
            continue
        cle = courriel or ("__stab" if nom == "Stab" else "")
        brut = MAIN.get(cle)
        txt = (brut.format(ANNONCE=ANNONCE, BENEFICE=BENEFICE, CLOTURE=CLOTURE)
               if brut else monter(prenom, nom, angle, region))
        mains += bool(brut)
        ws.cell(i, cB, txt).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(i, cO, OBJETS.get(angle, OBJETS["B"]))
        if not brut and salutation_douteuse(prenom, nom):
            cP = ent.index("Précaution") + 1 if "Précaution" in ent else None
            if cP:
                ancien = ws.cell(i, cP).value or ""
                ws.cell(i, cP, (ancien + " | " if ancien else "") +
                        "Salutation à vérifier : la fiche FPJQ est mal formée")
            douteux.append(f"{prenom} {nom}")
        ws.row_dimensions[i].height = 240
        par_angle.setdefault(angle, []).append((prio, prenom, nom, media, courriel, region,
                                                txt, bool(brut)))

    lignes = ["# Brouillons, liste froide FPJQ", "",
              "Même squelette que la liste chaude, mais l'ouverture part de ce qui concerne la "
              "personne : sa région quand l'asclépiade y a une histoire réelle, son sujet "
              "sinon.", "",
              "Salutation « Bonjour Prénom Nom » : deviner M. ou Mme sur 219 personnes qu'on "
              "ne connaît pas produirait des erreurs.", ""]
    for a in sorted(par_angle):
        lignes += [f"## Angle {a} — {len(par_angle[a])} contacts", ""]
        for prio, prenom, nom, media, courriel, region, txt, main_ in par_angle[a]:
            adr = f"`{courriel}`" if courriel else "*adresse à trouver*"
            tag = " — **écrit à la main**" if main_ else ""
            lignes += [f"### {prenom} {nom} — {media} — {adr}{tag}",
                       f"*Priorité {prio} · {region}*", "", "```", txt, "```", ""]

    wb.save(dst)
    open(md, "w").write("\n".join(lignes))
    total = sum(len(v) for v in par_angle.values())
    print(f"{total} brouillons froids réécrits, dont {mains} à la main")
    if douteux:
        print(f"  {len(douteux)} salutations signalées à vérifier : {', '.join(douteux[:6])}"
              + (" …" if len(douteux) > 6 else ""))
    for a in sorted(par_angle):
        print(f"  angle {a} : {len(par_angle[a])}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3])

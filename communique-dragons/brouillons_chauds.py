#!/usr/bin/env python3
"""Les 37 brouillons de la liste chaude, dans le registre de Gabriel.

Le squelette et les regles viennent de voix_gabriel.py, tire des deux courriels
que Gabriel a reecrits lui-meme. Ici, seul le contenu propre a chaque personne
est ecrit a la main : le rappel de son article ou de son echange, et le pont
vers la nouvelle.

Deux courriels sont repris mot pour mot de Gabriel : Larocque et Pouliot.

    python3 brouillons_chauds.py <chiffrier.xlsx> <sortie.xlsx> <sortie.md>
"""

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from voix_gabriel import (VIDEO, MISSIONS, ANNONCE, PAS, BENEFICE, OBJETS, MEDIA_KIT, deplier,
                          tu, cloture, assembler)

# --- ponts reutilisables ----------------------------------------------------

PONT_PIVOT = (
    "Parmi les raisons de ce changement, il y avait clairement que j'avais besoin de temps "
    f"pour me recentrer sur {MISSIONS}.\n\n{PAS}")

# « Depuis » et « Eh bien » supposent une histoire qui precede. Sans elle, la
# phrase tourne a vide : trois ponts distincts plutot qu'un seul mal etire.
PONT_MISSION = (
    f"Depuis, tout mon temps va à {MISSIONS}.\n\n{PAS}")

# Les quatre contacts qui recoivent la presentation complete y lisent deja la
# mission : la repeter ici serait redondant.
PONT_PRESENTATION = (
    "Le 17 septembre prochain, on va faire un énorme pas dans la bonne direction :")

PONT_DIRECT = PAS


def rarete(t, suite):
    """Le paragraphe de la rarete, avec la demande de suite adaptee."""
    base = ("Aller présenter notre entreprise et sa mission à la télévision nationale est une "
            "opportunité qui arrive bien rarement.")
    p, tp = ("t'", "ton") if t else ("vous ", "votre")
    if suite == "article":
        return f"{base} Je voulais {p}en faire part et qui sait, peut-être {p}inspirer une suite à {tp} dernier article."
    if suite == "entrevue":
        return (f"{base} Je voulais {p}en faire part : je suis disponible pour une entrevue la "
                "semaine du 14 septembre et le vendredi 18 au matin.")
    if suite == "sujet":
        return f"{base} Je voulais {p}en faire part et qui sait, peut-être {p}inspirer un sujet."
    return f"{base} Je voulais {p}en faire part."


# courriel : (prenom, nom, civilite, registre, rappel, pont, suite)
# Le registre par defaut est « vous » : Gabriel le bascule dans le chiffrier.
CHAUDS = {

# --- deux courriels ecrits par Gabriel lui-meme, repris tels quels -----------

"sylvain.larocque@quebecormedia.com": ("Sylvain", "Larocque", "M.", "vous", None, None, None,
"""Bonjour M. Larocque,

En mai dernier, nous avons discuté au téléphone suite à quoi, votre article sur le
déménagement de notre production textile en Tunisie a été publié et TRÈS relayé.

Je vous en remercie sincèrement. Le sujet était délicat et vous l'avez traité avec justesse :
de notre côté, le bilan de cet article est très positif.

Parmi mes projets avec tout le temps que j'allais sauver, il y avait de me concentrer sur
« faire connaître l'asclépiade » au plus grand nombre de gens possible. Eh bien, le
17 septembre prochain on va faire un énorme pas dans la bonne direction :

Je vais présenter Lasclay dans le premier épisode de la 21e saison de Dragons' Den, sur CBC
et CBC Gem.

Aller présenter notre entreprise et sa mission à la télévision nationale est une opportunité
qui arrive bien rarement. Je voulais vous en faire part et qui sait, peut-être vous inspirer
une suite au dernier article.

Notre média kit est ici, avec les images de notre passage à Dragons' Den et de l'entreprise :
https://drive.google.com/drive/folders/1pyCUbfHYQhpXXl4FoCC2RCFXKRvGS5Zr

J'espère vraiment un boom des ventes avec cette visibilité, ce
qui, plus largement, sera extrêmement bénéfique pour les cultivateurs d'asclépiade du Québec
chez qui on continue d'acheter, et pour les papillons monarques menacés qui continuent de se
reproduire dans leurs plantations.

Au plaisir et n'hésitez pas à me contacter si vous avez des questions."""),

"cpouliot@lesoleil.com": ("Chloé", "Pouliot", "Mme", "tu", None, None, None,
"""Bonjour Chloé,

J'espère que tu vas bien. Tu as peut-être vu passer mon vidéo portant sur notre changement
de modèle manufacturier (https://www.youtube.com/watch?v=GKyHh-Ok9JU) ou l'article de
Sylvain Larocque en mai dernier sur ce sujet. C'est en quelque sorte une suite logique à
l'article « Lasclay devant le dilemme de fabriquer au Québec » que tu avais publié en
décembre dernier.

Merci encore pour ce texte, d'ailleurs. Tu avais posé le dilemme sans le trancher à ma
place, et c'est exactement comme ça qu'il fallait le poser.

Neuf mois plus tard, le dilemme est tranché et il tient : l'asclépiade et l'isolant restent
à Limoilou, la coquille se fait en Tunisie, et le manteau se vend autour de 300 $.

Parmi les raisons de ce changement, il y avait clairement que j'avais besoin de temps pour
me recentrer sur mes 2 grandes missions : faire connaître l'asclépiade et sauvegarder les
monarques par la culture de l'asclépiade.

Eh bien, le 17 septembre prochain on va faire un énorme pas dans la bonne direction :

Je vais présenter Lasclay dans le premier épisode de la 21e saison de Dragons' Den, sur CBC
et CBC Gem.

Aller présenter notre entreprise et sa mission à la télévision nationale est une opportunité
qui arrive bien rarement. Je voulais t'en faire part et qui sait, peut-être t'inspirer une
suite à ton dernier article.

Notre média kit est ici, avec les images de notre passage à Dragons' Den et de l'entreprise :
https://drive.google.com/drive/folders/1pyCUbfHYQhpXXl4FoCC2RCFXKRvGS5Zr

J'espère vraiment un boom des ventes avec cette visibilité, ce
qui, plus largement, sera extrêmement bénéfique pour les cultivateurs d'asclépiade du Québec
chez qui on continue d'acheter, et pour les papillons monarques menacés qui continuent de se
reproduire dans leurs plantations.

Au plaisir et n'hésite pas à me contacter si tu as des questions."""),

# --- assembles sur le meme squelette ----------------------------------------

"sberube@lapresse.ca": ("Stéphanie", "Bérubé", "Mme", "vous",
 "En février, vous avez signé « La sinueuse route de la soie du Nord ». C'est un des rares "
 "textes à avoir pris la filière au complet, et pas juste une entreprise.",
 PONT_MISSION, "article", None),

"melissa.paquet@tva.ca": ("Melissa", "Paquet", "Mme", "vous",
 "On s'était parlé en septembre dernier pour Pleins feux sur Québec. Cette année, la date "
 "tombe presque au même endroit.",
 PONT_DIRECT, "entrevue", None),

"anne-sophie.roy@radio-canada.ca": ("Anne-Sophie", "Roy", "Mme", "vous",
 "Vous nous aviez appelés en mars 2025 pour votre reportage sur le dropshipping, comme "
 "exemple d'une entreprise qui fabrique vraiment. Et en décembre 2020, vous aviez signé un "
 "des tout premiers articles sur nos mitaines.",
 PONT_MISSION, "article", None),

"alafrance@lesoleil.com": ("Annie", "Lafrance", "Mme", "vous",
 "En avril 2024, vous aviez signé « L'asclépiade : plus que la fibre de demain » dans Le "
 "Soleil Affaires, où vous m'aviez cité disant qu'on croyait à cette matière depuis les "
 "débuts. On y croit encore.",
 PONT_PIVOT, "article", None),

"mtison@lapresse.ca": ("Marie", "Tison", "Mme", "vous",
 "En novembre 2021, vous aviez écrit « Le timide retour de l'asclépiade ». Le retour est "
 "pas mal moins timide qu'à l'époque : il y a maintenant un manteau et une veste à inserts "
 "d'asclépiade amovibles, autour de 300 $.",
 PONT_MISSION, "article", None),

"sylvie.lacombe@tva.ca": ("Sylvie", "Lacombe", "Mme", "vous",
 "En novembre 2023, vous nous aviez contactés pour le suivi d'une précommande qui traînait. "
 "Vous aviez raison de le faire. Depuis, on sort graduellement de la précommande pour vendre "
 "du déjà produit, et les délais n'ont plus rien à voir.",
 PONT_MISSION, "sujet", None),

"hrganzmann@ledevoir.com": ("Hélène", "Roulot-Ganzmann", "Mme", "vous",
 "On s'était parlé en novembre 2023 pour du matériel sur nos produits d'asclépiade.",
 PONT_MISSION, "sujet", None),

"msbrault@lesoleil.com": ("Marie-Soleil", "Brault", "Mme", "vous",
 "On s'était parlé en janvier 2023 pour Le Soleil. L'atelier est toujours dans Limoilou, "
 "et c'est encore là que la soie d'asclépiade devient de l'isolant.",
 PONT_MISSION, "article", None),

"sandra.fillion@radio-canada.ca": ("Sandra", "Fillion", "Mme", "vous",
 "On s'était parlé d'asclépiade en janvier 2023. Depuis, le statut du monarque comme espèce "
 "en voie de disparition a été confirmé au Canada, et les États-Unis ont proposé de "
 "l'inscrire comme menacée. La plante n'a jamais été aussi pertinente.",
 PONT_MISSION, "sujet", None),

"aabonn@latribune.qc.ca": ("Ariane", "Aubert Bonn", "Mme", "vous",
 "En novembre 2022, vous aviez couvert nos nouveaux produits d'asclépiade. Le catalogue "
 "compte maintenant plus de 40 produits, et La Tribune a repris en décembre dernier "
 "l'article de Chloé Pouliot sur notre changement de modèle manufacturier.",
 PONT_PIVOT, "article", None),

"eugenie.emond@radio-canada.ca": ("Eugénie", "Émond", "Mme", "vous",
 "En septembre 2022, on s'était parlé de notre glacière d'asclépiade imprimée en 3D. On a "
 "continué à inventer nos propres procédés depuis, parce que personne d'autre ne les fait.",
 PONT_MISSION, "sujet", None),

"rportelance@ledevoir.com": ("Rhéane", "Portelance", "Mme", "vous",
 "On avait échangé en juillet 2022 pour le cahier Action climatique. L'asclépiade est une "
 "vivace indigène qui se cultive avec peu d'intrants et qui stocke du carbone dans un "
 "système racinaire pérenne.",
 PONT_MISSION, "sujet", None),

"anais.elboujdaini@bellmedia.ca": ("Anaïs", "Elboujdaini", "Mme", "vous",
 "On s'était parlé du programme de semences en mai 2022. La Campagne nationale de plantation "
 "en est à sa 5e édition, et on a distribué environ 10 millions de graines d'asclépiade en "
 "Amérique du Nord.",
 PONT_DIRECT, "entrevue", None),

"mtison1@lapresse.ca": ("Marc", "Tison", "M.", "vous",
 "On s'était parlé pour la section Affaires en septembre 2021, à l'époque où on venait de "
 "rapatrier notre production faute de sous-traitant prêt à toucher à la fibre. Le modèle a "
 "changé depuis : la vidéo est ici, " + VIDEO + ", et Sylvain Larocque en a parlé en mai.",
 PONT_PIVOT, "article", None),

"vsimard@lapresse.ca": ("Valérie", "Simard", "Mme", "vous",
 "Vous aviez présenté nos moufles isolées à l'asclépiade le 17 septembre 2021. Drôle de "
 "coïncidence : la diffusion tombe cinq ans jour pour jour après votre article.",
 PONT_MISSION, "article", None),

"jfriis@unpointcinq.ca": ("Justine", "Friis", "Mme", "vous",
 "Vous aviez travaillé sur l'impact climatique de l'asclépiade en septembre 2021.",
 PONT_MISSION, "sujet", None),

"alex.perreault@radio-canada.ca": ("Alex", "Perreault", "M.", "vous",
 "On s'était parlé en juillet 2021.",
 PONT_MISSION, "entrevue", None),

"fhiggins@lesoleil.com": ("Francis", "Higgins", "M.", "vous",
 "On s'était parlé pour Le Soleil en juillet 2021. L'atelier est toujours dans Limoilou.",
 PONT_MISSION, "article", None),

"fanny.samson@radio-canada.ca": ("Fanny", "Samson", "Mme", "vous",
 "En mars 2021, vous aviez signé « Les grandes promesses de l'asclépiade, la soie "
 "d'Amérique ». Cinq ans plus tard, je peux vous dire lesquelles ont été tenues : il y a des "
 "produits finis, l'isolant se transforme toujours au Québec, on achète encore de "
 "l'asclépiade québécoise, et un manteau se vend autour de 300 $.",
 PONT_MISSION, "article", None),

"mmenard@laterre.ca": ("Martin", "Ménard", "M.", "vous",
 "On s'était parlé à La Terre en janvier 2021. Cinq ans plus tard, on achète encore de "
 "l'asclépiade québécoise et on la transforme nous-mêmes à Québec.",
 PONT_MISSION, "article", None),

"raphaelle.drouin@urbania.ca": ("Raphaëlle", "Drouin", "Mme", "vous",
 "En janvier 2021, vous aviez écrit « Devenir viral avant même de se lancer en affaires ». "
 "Six ans plus tard, l'entreprise existe pour vrai, avec plus de 40 produits.",
 PONT_MISSION, "article", None),

"groy@lequotidien.com": ("Guillaume", "Roy", "M.", "vous",
 "En octobre 2020, vous avez écrit « La deuxième vie de l'asclépiade ». Vous avez été le "
 "tout premier journaliste à parler de nous. Et l'agriculteur du Lac-Saint-Jean qui vous "
 "avait donné envie du sujet est encore notre fournisseur aujourd'hui : six ans plus tard, "
 "l'asclépiade de chez vous se retrouve toujours dans nos produits.",
 PONT_MISSION, "article", None),

"mireille.roberge@radio-canada.ca": ("Mireille", "Roberge", "Mme", "vous",
 "Je m'appelle Gabriel Gouveia, fondateur de Lasclay. On isole des vêtements d'hiver avec "
 "une mauvaise herbe, l'asclépiade, qu'on cultive pour sauvegarder un pollinisateur "
 "emblématique et menacé : le papillon monarque. Ses gousses sont remplies d'une soie "
 "creuse, très légère et naturellement hydrophobe, qu'on transforme en isolant à Québec.",
 PONT_PRESENTATION, "entrevue", None),

"sophie.laforest@radio-canada.ca": ("Sophie", "Laforest", "Mme", "vous",
 "Je m'appelle Gabriel Gouveia, fondateur de Lasclay. On isole des vêtements d'hiver avec "
 "une mauvaise herbe, l'asclépiade, qu'on cultive pour sauvegarder un pollinisateur "
 "emblématique et menacé : le papillon monarque. Ses gousses sont remplies d'une soie "
 "creuse, très légère et naturellement hydrophobe, qu'on transforme en isolant à Québec.",
 PONT_PRESENTATION, "entrevue", None),

"amckenna@ledevoir.com": ("Alain", "McKenna", "M.", "vous",
 "En décembre 2021, vous aviez écrit « Des écouteurs de Québec pour oublier les AirPods », "
 "sur Sounds Good. J'étais un des trois. J'ai vendu mes parts en 2022 pour me consacrer "
 "entièrement à Lasclay, l'autre entreprise que j'avais démarrée entretemps, qui transforme "
 "la soie d'asclépiade en isolant textile.",
 PONT_MISSION, "article", None),

"sylvielemieux16@gmail.com": ("Sylvie", "Lemieux", "Mme", "vous",
 "En décembre 2021, vous aviez écrit « Les deux mains dans l'écoresponsabilité » pour le "
 "Journal de Montréal. Notre modèle manufacturier a changé depuis, et j'en ai fait une vidéo "
 "ici : " + VIDEO + ".",
 PONT_PIVOT, "article", None),

"sophiegrenierheroux@hotmail.com": ("Sophie", "Grenier-Héroux", "Mme", "vous",
 "En janvier 2021, vous aviez écrit « La soie d'Amérique, le pari de deux ambitieux » dans "
 "Le Devoir. Le pari a six ans, et il tient : plus de 40 produits, environ 10 millions de "
 "graines distribuées, et un isolant toujours transformé à Limoilou.",
 PONT_MISSION, "article", None),

"madeleine.goubau@gmail.com": ("Madeleine", "Goubau", "Mme", "vous",
 "Je m'appelle Gabriel Gouveia, fondateur de Lasclay. On isole des vêtements d'hiver avec "
 "une mauvaise herbe : l'asclépiade, que vous connaissez bien, puisque vous en aviez parlé à "
 "Moteur de recherche en février 2023. On la cultive pour sauvegarder un pollinisateur "
 "emblématique et menacé, le papillon monarque.\n\nCe que vous n'avez peut-être pas vu de "
 "près, c'est ce que la fibre donne une fois transformée. Ses gousses sont remplies d'une "
 "soie creuse, très légère et naturellement hydrophobe, qu'on transforme en isolant à "
 "Québec, et qui se retrouve dans des manteaux, des mitaines et des tuques.",
 PONT_PRESENTATION, "sujet", None),

"jessica.dostie@gmail.com": ("Jessica", "Dostie", "Mme", "vous",
 "Je m'appelle Gabriel Gouveia, fondateur de Lasclay. On isole des vêtements d'hiver avec "
 "une mauvaise herbe, l'asclépiade, qu'on cultive pour sauvegarder un pollinisateur "
 "emblématique et menacé : le papillon monarque. Ses gousses sont remplies d'une soie "
 "creuse, très légère et naturellement hydrophobe, qu'on transforme en isolant à Québec.",
 PONT_PRESENTATION, "sujet", None),

"__poisson": ("Sophie", "Poisson", "Mme", "vous",
 "En novembre 2020, vous aviez présenté nos accessoires d'hiver isolés à l'asclépiade, quand "
 "il y en avait trois. Il y en a maintenant plus de 40, dont des manteaux et des vestes à "
 "inserts amovibles.",
 PONT_MISSION, "article", None),

"__bertrand": ("Caroline", "Bertrand", "Mme", "vous",
 "En septembre 2021, vous aviez écrit « L'asclépiade pour braver le froid » pour ICI "
 "Explora. Depuis, le statut du monarque comme espèce en voie de disparition a été confirmé "
 "au Canada, et les États-Unis ont proposé de l'inscrire comme menacée.",
 PONT_MISSION, "article", None),

"__benoist": ("Karine", "Benoist", "Mme", "vous",
 "En novembre 2023, vous aviez retenu notre foulard d'asclépiade dans vos cadeaux faits au "
 "Québec. Une précision utile pour la prochaine fois : notre modèle manufacturier a changé "
 "en 2025, et l'assemblage des produits finis se fait maintenant en Tunisie. L'asclépiade "
 "est cultivée au Québec et l'isolant transformé à Limoilou.",
 PONT_PIVOT, "article", None),
}

# --- le remerciement ---------------------------------------------------------
# Gabriel : « Remercie chaque media de la liste chaude pour leur article. »
# Une seule regle, et elle est severe : ne remercier que pour ce qu'on sait
# avoir ete publie, et dire ce que le texte a fait de bien, jamais « merci pour
# votre article » tout court. Trois contacts n'ont aucun historique connu, ils
# n'ont donc pas de remerciement : on ne remercie pas pour un article
# hypothetique.

MERCIS = {
"sberube@lapresse.ca":
    "Merci d'avoir pris cette peine-là. La filière a plus besoin d'être racontée en entier "
    "qu'entreprise par entreprise, et presque personne ne le fait.",

"melissa.paquet@tva.ca":
    "Merci pour la tribune de l'an dernier. Ce genre de rendez-vous local vaut cher pour une "
    "entreprise de notre taille.",

"anne-sophie.roy@radio-canada.ca":
    "Merci pour les deux. Être cité comme l'exemple d'une entreprise qui fabrique vraiment, "
    "dans un reportage sur le dropshipping, c'est exactement la distinction qu'on essaie de "
    "faire comprendre.",

"alafrance@lesoleil.com":
    "Merci d'avoir gardé cette citation-là : elle a bien vieilli.",

"mtison@lapresse.ca":
    "Merci d'avoir écrit sur la plante à une époque où il n'y avait presque rien à montrer. "
    "C'est le genre de texte qui aide une filière avant qu'elle existe.",

"sylvie.lacombe@tva.ca":
    "Merci de nous l'avoir signalé plutôt que de laisser passer : c'est comme ça qu'on a "
    "corrigé le tir.",

"hrganzmann@ledevoir.com":
    "Merci de vous être intéressée à la matière elle-même. C'est rarement le premier angle "
    "qu'on nous propose.",

"msbrault@lesoleil.com":
    "Merci pour l'attention que vous nous aviez portée à l'époque : Le Soleil a été un des "
    "premiers à nous suivre sérieusement.",

"sandra.fillion@radio-canada.ca":
    "Merci d'avoir pris l'asclépiade au sérieux à un moment où bien peu de monde le faisait.",

"aabonn@latribune.qc.ca":
    "Merci de l'avoir fait quand le catalogue tenait encore sur une main. L'Estrie nous suit "
    "depuis, et ça part de là.",

"eugenie.emond@radio-canada.ca":
    "Merci de vous être arrêtée sur un détail technique. C'est rare, et c'est pourtant là que "
    "se joue la différence entre une vraie matière et un argument de vente.",

"rportelance@ledevoir.com":
    "Merci d'avoir ouvert le cahier climatique à une plante : les solutions agricoles y "
    "passent moins souvent que les technologies.",

"anais.elboujdaini@bellmedia.ca":
    "Merci d'avoir parlé du programme de semences. C'est la partie la moins spectaculaire de "
    "ce qu'on fait, et la plus utile aux monarques.",

"mtison1@lapresse.ca":
    "Merci d'avoir raconté ce rapatriement au moment où c'était encore un pari. Le texte a "
    "beaucoup circulé chez nos clients.",

"vsimard@lapresse.ca":
    "Merci pour ce coup de pouce de la première heure : à l'époque, une mention dans La "
    "Presse changeait notre semaine.",

"jfriis@unpointcinq.ca":
    "Merci de vous être penchée sur l'impact réel plutôt que sur l'intention. C'est ce qui "
    "sépare une plante utile d'un argument vert.",

"alex.perreault@radio-canada.ca":
    "Merci de l'intérêt que vous nous aviez porté à l'époque : on partait de zéro.",

"fhiggins@lesoleil.com":
    "Merci d'avoir couvert l'atelier quand il n'était qu'un local et deux machines.",

"fanny.samson@radio-canada.ca":
    "Merci d'avoir écrit « promesses » plutôt qu'autre chose. Le mot était juste, et il l'est "
    "resté le temps qu'on les tienne.",

"mmenard@laterre.ca":
    "Merci d'avoir porté le sujet auprès des producteurs. C'est le lectorat qui compte le plus "
    "pour nous : sans champs, il n'y a pas de fibre.",

"raphaelle.drouin@urbania.ca":
    "Merci d'avoir écrit sur nous quand il n'y avait littéralement rien à acheter. Vous nous "
    "avez pris au sérieux avant tout le monde.",

"groy@lequotidien.com":
    "Merci, sincèrement. On n'oublie pas qui a écrit le premier article, et le vôtre nous a "
    "servi de carte de visite pendant des années.",

"amckenna@ledevoir.com":
    "Merci pour ce texte-là : il a compté pour Sounds Good, et j'en garde un bon souvenir.",

"sylvielemieux16@gmail.com":
    "Merci d'avoir mis l'écoresponsabilité au cœur du texte plutôt qu'en décoration.",

"sophiegrenierheroux@hotmail.com":
    "Merci d'avoir appelé ça un pari. C'en était un, et le mot était plus honnête que tous "
    "les superlatifs qu'on nous colle habituellement.",

"madeleine.goubau@gmail.com":
    "Merci d'en avoir parlé en ondes, d'ailleurs : c'est rare qu'on explique la plante avant "
    "de vendre quoi que ce soit.",

"__poisson":
    "Merci d'avoir montré les trois premiers produits : c'était notre première couverture "
    "dans un magazine de design.",

"__bertrand":
    "Merci d'avoir fait le lien entre le froid et le monarque. C'est le raccourci le plus "
    "difficile à faire passer, et vous l'aviez fait en un titre.",

"__benoist":
    "Merci de nous avoir retenus dans cette sélection-là : les cadeaux faits au Québec, c'est "
    "une vitrine qui compte pour un petit atelier.",
}

# Retires de la liste chaude sur demande de Gabriel : leurs lignes sont
# supprimees du chiffrier, pas seulement marquees.
EXCLUS = {"jhaurio@unpointcinq.ca", "contact@protegez-vous.ca"}

# Sans adresse courriel : Sophie Poisson (Baron Mag), Caroline Bertrand (ICI
# Explora) et Karine Benoist (Chatelaine). Leurs brouillons restent dans
# CHAUDS si leurs adresses refont surface, mais elles sortent du chiffrier.
EXCLUS_SANS_ADRESSE = {"Sophie Poisson", "Caroline Bertrand", "Karine Benoist"}

NE_PAS_ENVOYER = {
"anne-sophie.roy@quebecormedia.com":
    "NE PAS ENVOYER — même journaliste que la ligne Radio-Canada, à une ancienne adresse. "
    "Écrire à anne-sophie.roy@radio-canada.ca. Garder cette adresse en réserve si l'autre "
    "rebondit.",
}


def monter(cle, registre=None):
    prenom, nom, civ, reg_defaut, rappel, pont, suite, brut = (
        CHAUDS[cle] + (None,) * 8)[:8]
    reg = (registre or reg_defaut or "vous").strip().lower()
    if brut:
        return deplier(brut)
    t = tu(reg)
    if prenom:
        salut = f"Bonjour {prenom}," if t else f"Bonjour {civ} {nom},"
    else:
        salut = "Bonjour,"
    tete = salut + ("\n\nJ'espère que tu vas bien." if t else "")
    return assembler([tete, rappel, MERCIS.get(cle), pont, ANNONCE, rarete(t, suite),
                      MEDIA_KIT, BENEFICE, cloture(reg)])


def main(src, dst, md):
    wb = openpyxl.load_workbook(src)
    ws = next(w for w in wb.worksheets if w.title.startswith("Liste chaude"))
    ent = [c.value for c in ws[1]]
    cO = ent.index("Objet suggéré") + 1
    cAng = ent.index("Angle") + 1
    cB = ent.index("Brouillon") + 1 if "Brouillon" in ent else ws.max_column + 1
    cR = ent.index("Registre") + 1 if "Registre" in ent else ws.max_column + (
        1 if "Brouillon" in ent else 2)
    for j, titre in ((cB, "Brouillon"), (cR, "Registre")):
        c = ws.cell(1, j, titre)
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="FFF2CC" if titre == "Registre" else "E3EADF")
    ws.column_dimensions[ws.cell(1, cB).column_letter].width = 78
    ws.column_dimensions[ws.cell(1, cR).column_letter].width = 12

    a_supprimer = []
    lignes, faits, manquants = ["# Brouillons, liste chaude", "",
        "Registre de Gabriel, tiré des deux courriels qu'il a réécrits lui-même. Larocque et "
        "Pouliot sont repris mot pour mot.", "",
        "La colonne **Registre** du chiffrier tranche entre `tu` et `vous`. Elle ne se devine "
        "pas : bascule-la, et relance le script.", ""], 0, []

    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not r[0]:
            continue
        courriel = (r[2] or "").strip().lower()
        nom = (r[0] or "").strip()
        if courriel in EXCLUS or (not courriel and nom in EXCLUS_SANS_ADRESSE):
            a_supprimer.append(i)
            continue
        cle = courriel
        if not courriel:
            cle = {"Sophie Poisson": "__poisson", "Caroline Bertrand": "__bertrand",
                   "Karine Benoist": "__benoist"}.get(nom, "")
        if courriel in NE_PAS_ENVOYER:
            ws.cell(i, cB, NE_PAS_ENVOYER[courriel]).alignment = Alignment(
                wrap_text=True, vertical="top")
            ws.cell(i, cO, "—")
            ws.cell(i, cR, "—")
            continue
        if cle not in CHAUDS:
            manquants.append(f"{nom} <{courriel}>")
            continue
        reg = (r[cR - 1] if cR - 1 < len(r) else None) or CHAUDS[cle][3]
        txt = monter(cle, reg)
        ws.cell(i, cB, txt).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(i, cO, OBJETS.get(r[cAng - 1] or "A", OBJETS["A"]))
        ws.cell(i, cR, reg)
        ws.row_dimensions[i].height = 240
        faits += 1
        adr = f"`{courriel}`" if courriel else "*adresse à trouver*"
        lignes += [f"## {nom} — {r[1]} — {adr}", f"*Registre : {reg}*", "", "```", txt, "```", ""]

    for i in reversed(a_supprimer):
        ws.delete_rows(i)

    wb.save(dst)
    open(md, "w").write("\n".join(lignes))
    if a_supprimer:
        print(f"{len(a_supprimer)} lignes retirées de la liste chaude")
    print(f"{faits} brouillons chauds réécrits dans {dst} et {md}")
    if manquants:
        print("sans texte :", manquants)


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3])

#!/usr/bin/env python3
"""Assigne un angle de personnalisation a chacun des 251 contacts des listes medias.

Lit le chiffrier fourni, ajoute deux colonnes (Angle, Objet suggere) et un onglet
"Plan d'envoi" qui croise les vagues du chiffrier avec les angles de la matrice.
Les colonnes de donnees recoltees et les colonnes jaunes (Envoye le, Reponse,
Notes) ne sont jamais touchees.

    python3 assigner_angles.py <source.xlsx> <destination.xlsx>
"""

import sys
import collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# --- les neuf angles de 03-matrice-personnalisation.md -----------------------

OBJETS = {
    "A": "Suite a votre article de {annee} sur l'asclepiade",
    "B": "{region} : Lasclay au premier episode de la 21e saison de Dragons' Den",
    "C": "L'asclepiade, huit ans apres l'effondrement de la filiere",
    "D": "Le monarque, l'asclepiade, et l'idee qu'il faut la rendre payante",
    "E": "Une PME qui a rapatrie sa production, puis en a delocalise une partie",
    "F": "L'asclepiade contre le duvet, et ce que les tests disent vraiment",
    "G": "Un manteau isole avec une mauvaise herbe, au premier episode de Dragons' Den",
    "H": "The weed that insulates a winter coat, on the season 21 premiere",
    "I": "Invite disponible : Lasclay a Dragons' Den le 17 septembre",
    # Angle ne du depouillement du web : des journalistes ont couvert
    # l'asclepiade en profondeur sans jamais avoir parle a Lasclay. Ce sont les
    # meilleurs contacts froids du dossier, et le seul angle qui leur parle est
    # leur propre couverture de la filiere.
    "J": "Vous couvrez l'asclepiade depuis {annee}. Il y a une suite.",
}

# --- ajouts issus du depouillement du web -----------------------------------
# Regle absolue : aucune adresse n'est devinee. Un courriel introuvable reste
# vide et part en verification manuelle.

AJOUTS_CHAUDS = [
    # nom, media, courriel, date, sujet, angle, note
    ("Sophie Poisson", "Baron Mag", "", "2020-11-10",
     "Lasclay : des accessoires d'hiver isoles a l'asclepiade", "A",
     "A mentionne Lasclay. Absente des 34. Courriel a trouver."),
    ("Caroline Bertrand", "ICI Explora", "", "2021-09-28",
     "L'asclepiade pour braver le froid", "A",
     "A mentionne Lasclay. Absente des 34. Courriel a trouver."),
    ("Karine Benoist", "Chatelaine", "", "2023-11-29",
     "Cadeaux faits au Quebec : foulard d'asclepiade", "A",
     "A mentionne Lasclay. Absente des 34. Courriel a trouver."),
]

AJOUTS_FROIDS = [
    # prio, prenom, nom, fonction, media, courriel, region, secteurs, angle, note
    ("A", "Jean-Michel", "Leprince", "Journaliste", "Radio-Canada",
     "jean-michel_leprince@radio-canada.ca", "Montreal",
     "Agriculture; Environnement; Sciences", "J",
     "Nombreux reportages sur l'asclepiade au Telejournal depuis 2014. "
     "Jamais de contact avec Lasclay. Adresse fournie par Gabriel."),
    ("B", "Antoine", "Stab", "Journaliste", "Espaces", "", "Montreal",
     "Environnement; Plein air", "J",
     "A signe sur le soyer du Quebec dans Espaces (2015-02-16). "
     "Courriel a trouver, et verifier qu'il y est toujours."),
]

# Presents dans les deux listes. Ils appartiennent a la liste chaude : le
# message ecrit a la main l'emporte sur la sollicitation froide.
DOUBLONS = {"alafrance@lesoleil.com", "fhiggins@lesoleil.com"}

# Signatures resolues par le depouillement, qui etaient marquees a verifier.
SIGNATURES_RESOLUES = {
    "Le Soleil Affaires, 2024-04-20, L'asclepiade : plus que la fibre de demain":
        "Annie Lafrance (confirme)",
    "Le Soleil / La Tribune / Le Droit, 2025-12-01, Lasclay devant le dilemme":
        "Chloe Pouliot (confirme)",
    "La Presse, 2026-02-23, La sinueuse route de la soie du Nord":
        "Stephanie Berube (probable : seul contact de la liste chaude a cette date exacte, a confirmer)",
}

# Medias nationaux : la proximite geographique ne joue pas, on garde le thematique.
NATIONAUX = {
    "la presse", "le devoir", "les affaires", "protegez-vous", "protégez-vous",
    "la terre de chez nous", "urbania", "unpointcinq", "noovo info", "quebecor media",
    "quebecor média", "bell media", "bell média", "radio-canada (montreal)",
    "radio-canada (montréal)", "radio-canada (radio - montreal)",
    "radio-canada (radio - montréal)", "l'actualite", "l'actualité", "ricochet",
    "la conversation canada", "quebec science", "québec science",
}

# Attention : "Radio-Canada" contient "radio". Ne detecter que les antennes
# reelles, sinon toute la salle de nouvelles ecrite bascule en angle I.
MOTS_RADIO_TV = (
    "(radio", "rdi", "tva", "noovo", "tfo", "tvontario", "radio vm",
    "ici premiere", "ici première", "cogeco", "98,5",
)

MOTS_AGRICOLE = ("terre de chez nous", "bulletin des agriculteurs", "agricole", "coop")
MOTS_AFFAIRES = ("affaires", "economi", "économi", "finance", "investissement")
MOTS_PLEINAIR = ("plein air", "geo plein air", "géo plein air", "espaces", "rando",
                 "chasse", "peche", "pêche", "velo", "vélo")
MOTS_ANGLAIS = ("gazette", "globe and mail", "national post", "financial post",
                "canadian press", "betakit", "maclean", "cbc news")

# Piege verifie sur les donnees : dans le repertoire FPJQ, la region
# "Canada anglais" designe la francophonie hors Quebec, pas la presse
# anglophone. Les sept fiches concernees sont Radio-Canada, TFO et Le Devoir.
# Aucune ne recoit l'angle H : elles restent francophones.
MEDIAS_FRANCO = ("radio-canada", "tfo", "tvontario", "le devoir", "onfr")


def secteurs(cell):
    return {s.strip().lower() for s in (cell or "").split(";") if s.strip()}


def angle_froid(media, region, fonction, secs):
    """Cascade du plus specifique au plus general. Ordre volontaire.

    Environnement couvre 127 des 217 fiches : le placer tot ecraserait tout le
    reste. Les angles rares passent donc devant.
    """
    m = (media or "").lower()
    r = (region or "").lower()
    f = (fonction or "").lower()

    # H : presse anglophone. Le contexte de 2018 ne veut rien dire pour elle.
    if any(w in m for w in MOTS_ANGLAIS) and not any(w in m for w in MEDIAS_FRANCO):
        return "H"

    # I : radio et tele. Ils veulent un invite, pas un communique.
    if any(w in m for w in MOTS_RADIO_TV) or "anima" in f or "recherchiste" in f:
        return "I"

    # B : media local hors Montreal. La proximite bat la thematique.
    if r and r != "montreal" and r != "montréal" and m != "pigiste" \
            and m not in NATIONAUX:
        return "B"

    # C : filiere agricole. Passe devant l'environnement volontairement.
    # L'effondrement de 2018 est le recit le moins generique de Lasclay, et
    # tout journaliste qui couvre l'agriculture le trouvera plus pertinent
    # qu'un enieme argumentaire sur le monarque.
    if any(w in m for w in MOTS_AGRICOLE) or "agriculture" in secs:
        return "C"

    # F : equipement et plein air.
    if any(w in m for w in MOTS_PLEINAIR) or "plein air" in secs:
        return "F"

    # E : affaires et manufacturier.
    if any(w in m for w in MOTS_AFFAIRES) or (
            "économie" in secs and not (secs & {"environnement", "sciences"})):
        return "E"

    # G : mode et beaute, un signal fort quand il est present.
    if "mode et beauté" in secs:
        return "G"

    # D : environnement et sciences, le gros du repertoire.
    if secs & {"environnement", "sciences"}:
        return "D"

    # G : art de vivre et consommation.
    if secs & {"art de vivre", "consommation"}:
        return "G"

    # E puis C en rattrapage, sinon D par defaut.
    if "économie" in secs:
        return "E"
    if "agriculture" in secs:
        return "C"
    return "D"


def angle_chaud(media, sujet):
    """Liste chaude : angle A par defaut, la regle de priorite de la matrice.

    Sauf pour la radio et la tele, ou le format commande, et pour le contact
    general d'un magazine, qui n'a pas d'article a lui.
    """
    m = (media or "").lower()
    if "contact general" in (media or "").lower():
        return "G"
    if any(w in m for w in ("tva", "bell media", "bell média")):
        return "I"
    return "A"


def perso(nom, prenom=None):
    """Prenom utilisable pour la salutation, ou rien si on ne le sait pas."""
    if prenom:
        return prenom.strip()
    parts = (nom or "").strip().split()
    if not parts:
        return ""
    # "G. Roy" ou "A. McKenna" : initiale seule, on ne devine pas.
    return "" if len(parts[0]) <= 2 or parts[0].endswith(".") else parts[0]


def main(src, dst):
    wb = openpyxl.load_workbook(src)
    tally = collections.Counter()
    manuels, ajouts_c, ajouts_f, doublons_vus = [], [], [], []
    derniere_ligne_chaude = derniere_ligne_froide = 1

    entete = Font(bold=True, size=10)
    fond = PatternFill("solid", fgColor="E3EADF")

    # --- liste chaude --------------------------------------------------------
    ws = next(w for w in wb.worksheets if w.title.startswith("Liste chaude"))
    col = ws.max_column + 1
    ws.cell(1, col, "Angle").font = entete
    ws.cell(1, col).fill = fond
    ws.cell(1, col + 1, "Objet suggéré").font = entete
    ws.cell(1, col + 1).fill = fond
    ws.cell(1, col + 2, "Salutation").font = entete
    ws.cell(1, col + 2).fill = fond

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        nom, media, courriel, dernier, sujet = row[0], row[1], row[2], row[3], row[4]
        a = angle_chaud(media, sujet)
        annee = str(dernier)[:4] if dernier else "[année]"
        ws.cell(i, col, a)
        ws.cell(i, col + 1, OBJETS[a].format(annee=annee, region=""))
        p = perso(nom)
        ws.cell(i, col + 2, f"Bonjour {p}," if p else "Bonjour,")
        tally["chaude " + a] += 1
        if a == "A":
            manuels.append((nom, media, dernier, sujet))
        derniere_ligne_chaude = i

    # Journalistes qui ont mentionne Lasclay mais qui manquaient a la liste.
    for nom, media, courriel, date, sujet, a, note in AJOUTS_CHAUDS:
        derniere_ligne_chaude += 1
        i = derniere_ligne_chaude
        ws.cell(i, 1, nom)
        ws.cell(i, 2, media)
        ws.cell(i, 3, courriel)
        ws.cell(i, 4, date)
        ws.cell(i, 5, sujet)
        ws.cell(i, 6, "Depouillement web (page medias de lasclay.com)")
        ws.cell(i, 7, "A verifier" if not courriel else "Correspondance confirmee")
        ws.cell(i, 10, note)
        ws.cell(i, col, a)
        ws.cell(i, col + 1, OBJETS[a].format(annee=date[:4], region=""))
        ws.cell(i, col + 2, f"Bonjour {nom.split()[0]},")
        tally["chaude " + a] += 1
        ajouts_c.append((nom, media, courriel))

    # --- liste froide --------------------------------------------------------
    wf = next(w for w in wb.worksheets if w.title.startswith("Liste froide"))
    col = wf.max_column + 1
    for j, titre in enumerate(("Angle", "Objet suggéré", "Salutation", "Précaution")):
        c = wf.cell(1, col + j, titre)
        c.font = entete
        c.fill = fond

    for i, row in enumerate(wf.iter_rows(min_row=2, values_only=True), start=2):
        if not row[1]:
            continue
        prio, prenom, nom, fonction, media, courriel = row[0:6]
        region, secs_raw = row[7], row[8]
        secs = secteurs(secs_raw)
        a = angle_froid(media, region, fonction, secs)
        wf.cell(i, col, a)
        wf.cell(i, col + 1, OBJETS[a].format(annee="", region=region or ""))
        wf.cell(i, col + 2, f"Bonjour {prenom.strip()}," if prenom else "Bonjour,")

        # LCAP : adresse personnelle de pigiste, message vraiment individualise.
        dom = (courriel or "").split("@")[-1].lower()
        if (media or "").strip().lower() == "pigiste" or dom in (
                "gmail.com", "hotmail.com", "videotron.ca", "outlook.com",
                "yahoo.ca", "yahoo.com", "sympatico.ca", "me.com", "icloud.com"):
            wf.cell(i, col + 3, "Adresse personnelle : individualiser vraiment")
        # Present aussi dans la liste chaude : le message ecrit a la main
        # l'emporte. On neutralise la ligne froide au lieu de la supprimer,
        # pour que le chiffrier de Gabriel garde sa tracabilite.
        if (courriel or "").strip().lower() in DOUBLONS:
            wf.cell(i, col, "—")
            wf.cell(i, col + 1, "")
            wf.cell(i, col + 3, "DOUBLON : deja traite en liste chaude, ne pas envoyer")
            doublons_vus.append(f"{prenom} {nom} ({courriel})")
            continue
        tally[f"froide {prio} {a}"] += 1
        tally["ANGLE " + a] += 1
        derniere_ligne_froide = i

    for prio, prenom, nom, fonction, media, courriel, region, secs_raw, a, note in AJOUTS_FROIDS:
        derniere_ligne_froide += 1
        i = derniere_ligne_froide
        for cidx, val in ((1, prio), (2, prenom), (3, nom), (4, fonction),
                          (5, media), (6, courriel), (8, region), (9, secs_raw)):
            wf.cell(i, cidx, val)
        wf.cell(i, 14, note)
        wf.cell(i, col, a)
        wf.cell(i, col + 1, OBJETS[a].format(annee="2014" if nom == "Leprince" else "2015",
                                             region=region))
        wf.cell(i, col + 2, f"Bonjour {prenom},")
        if not courriel:
            wf.cell(i, col + 3, "Courriel introuvable : a chercher avant l'envoi")
        tally[f"froide {prio} {a}"] += 1
        tally["ANGLE " + a] += 1
        ajouts_f.append((f"{prenom} {nom}", media, courriel))

    # --- onglet plan ---------------------------------------------------------
    if "Plan d'envoi" in wb.sheetnames:
        del wb["Plan d'envoi"]
    p = wb.create_sheet("Plan d'envoi", 3)
    p.column_dimensions["A"].width = 30
    p.column_dimensions["B"].width = 12
    p.column_dimensions["C"].width = 62

    def ligne(r, a, b="", c="", bold=False):
        p.cell(r, 1, a).font = Font(bold=bold or not b, size=11 if bold else 10)
        if b:
            p.cell(r, 2, b)
        if c:
            p.cell(r, 3, c).alignment = Alignment(wrap_text=True, vertical="top")

    r = 1
    ligne(r, "Répartition des angles", bold=True); r += 2
    ligne(r, "ANGLE", "CONTACTS", "CE QUE LE COURRIEL DOIT FAIRE"); r += 1
    quoi = {
        "A": "Citer son article par titre et année dès la première phrase, puis dire ce qui a changé, y compris ce qui a mal tourné.",
        "B": "Nommer la région dans l'objet. Un pupitre régional trie par toponyme.",
        "C": "Poser la vraie question : y a-t-il enfin un acheteur stable au bout du champ. Ne rien promettre en volumes.",
        "D": "Conservation par le débouché économique. Dire soi-même qu'un achat ne sauve pas un papillon.",
        "E": "Rapatrier puis délocaliser en partie. Aucun chiffre financier sans autorisation de Gabriel.",
        "F": "Offrir de tester. Le 10 % plus isolant que le duvet est un repère de laboratoire, jamais une promesse.",
        "G": "Le contraste visuel : gousse, soie, manteau porté. Jamais « fabriqué au Québec » sur un produit fini.",
        "H": "Commencer par la plante, pas par le Québec. Lasclay arrive en troisième phrase.",
        "I": "Trois lignes. Disponibilité en clair. Ils veulent un invité, pas un communiqué.",
        "J": "Citer sa propre couverture de la filière, pas celle de Lasclay. Il connaît le dossier mieux que la moyenne : ne rien lui expliquer qu'il sait déjà.",
    }
    for a in "ABCDEFGHIJ":
        n = tally["ANGLE " + a] + sum(v for k, v in tally.items()
                                      if k.startswith("chaude ") and k.endswith(a))
        ligne(r, f"Angle {a}", n, quoi[a]); r += 1
    r += 1

    ligne(r, "Vagues du chiffrier", bold=True); r += 2
    for a, b, c in (
        ("Maintenant → 2 sept.", "34", "Liste chaude. Écrits à la main, un par un. Angle A sauf exceptions."),
        ("3 → 5 sept.", "71", "Froide A. Environ 60 messages par jour, première ligne réellement différente."),
        ("8 → 9 sept.", "53", "Froide B."),
        ("12 sept., 9 h", "—", "Ouverture de la prévente. Klaviyo vers les clients, jamais vers ces listes."),
        ("15 → 16 sept.", "93", "Froide C, seulement s'il reste du souffle."),
        ("17 sept., 20 h", "—", "Diffusion. Relance le 18 au matin : la confidentialité tombe, le sujet devient vérifiable."),
    ):
        ligne(r, a, b, c); r += 1
    r += 1

    ligne(r, "Règles d'envoi", bold=True); r += 2
    for a, c in (
        ("Canal", "Missive depuis media@lasclay.com, un message à la fois. Jamais Klaviyo : des rebonds sur adresses froides dégraderaient mail.lasclay.com, le domaine qui doit livrer l'infolettre de prévente."),
        ("Suivi", "Désactiver le suivi des ouvertures et des clics. Le pixel n'apporte rien et plusieurs journalistes le voient mal."),
        ("Rythme", "Environ 60 par jour. 217 messages quasi identiques en dix minutes se comportent comme du publipostage aux yeux des filtres."),
        ("LCAP", "Sollicitation de presse ciblée vers une adresse professionnelle publiée : couvert. Les adresses personnelles de pigistes exigent un message vraiment individualisé."),
        ("Interdit", "Ajouter ces contacts à une liste de diffusion, leur envoyer une infolettre, ou écrire sans objet journalistique réel."),
        ("CBC", "Rien sur l'issue avant le 17. Aucun logo ni « vu à Dragons' Den ». alicia.chirrey@cbc.ca et kylee.habrowski@cbc.ca sont la production, pas la salle de nouvelles : ne pas les solliciter."),
    ):
        ligne(r, a, "", c); r += 1

    wb.save(dst)

    print(f"écrit : {dst}\n")
    print("Répartition des angles sur les 251 contacts")
    for a in "ABCDEFGHIJ":
        n = tally["ANGLE " + a] + sum(v for k, v in tally.items()
                                      if k.startswith("chaude ") and k.endswith(a))
        print(f"  angle {a}  {n:>4}")
    print("\nListe chaude")
    for k in sorted(k for k in tally if k.startswith("chaude ")):
        print(f"  {k}  {tally[k]}")
    print("\nListe froide, par priorité")
    for k in sorted(k for k in tally if k.startswith("froide ")):
        print(f"  {k}  {tally[k]}")
    print(f"\n{len(manuels)} messages écrits à la main (angle A de la liste chaude)")
    print(f"\nAjoutés à la liste chaude ({len(ajouts_c)}) — ont mentionné Lasclay :")
    for n, m, c in ajouts_c:
        print(f"  {n:<22} {m:<14} {c or 'COURRIEL À TROUVER'}")
    print(f"\nAjoutés à la liste froide ({len(ajouts_f)}) — ont couvert l'asclépiade :")
    for n, m, c in ajouts_f:
        print(f"  {n:<22} {m:<14} {c or 'COURRIEL À TROUVER'}")
    print(f"\nDoublons neutralisés dans la liste froide ({len(doublons_vus)}) :")
    for d in doublons_vus:
        print(f"  {d}")
    print("\nSignatures résolues par le dépouillement :")
    for k, v in SIGNATURES_RESOLUES.items():
        print(f"  {k}\n      -> {v}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])

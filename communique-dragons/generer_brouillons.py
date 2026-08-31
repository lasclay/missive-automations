#!/usr/bin/env python3
"""Genere un brouillon de courriel pour chaque contact de la liste froide.

Deux regimes. Les contacts a forte valeur ont un texte ecrit a la main, dans
ECRITS_A_LA_MAIN. Tous les autres passent par un assemblage: une accroche
regionale ou thematique reellement differente d'un contact a l'autre, puis le
corps de l'angle assigne.

Le chiffrier exige "une premiere ligne reellement differente pour chacun".
C'est exactement ce que fait l'accroche: elle vient de la region du contact
quand il y en a une qui mord, de son secteur sinon.

    python3 generer_brouillons.py <chiffrier_angles.xlsx> <sortie.xlsx> <sortie.md>
"""

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DIFFUSION = ("le 17 septembre, Lasclay est dans le premier épisode de la 21e "
             "saison de Dragons' Den, sur CBC et CBC Gem")

# --- accroches regionales ---------------------------------------------------
# Chaque region ou l'asclepiade a une histoire reelle a raconter. Ailleurs, on
# ne force pas un lien local qui n'existe pas.

REGIONS = {
    "Mauricie":
        "L'usine qui achetait 90 % des récoltes d'asclépiade du Québec était à Saint-Tite. "
        "Quand Les Industries Encore 3 ont fait faillite, la filière s'est effondrée et des "
        "producteurs de la Mauricie ont retourné leurs champs.",
    "Centre-du-Québec":
        "L'asclépiade se cultive dans votre région depuis la première vague de 2013, et "
        "plusieurs producteurs qui ont survécu à l'effondrement de la filière y sont encore.",
    "Estrie":
        "Une des usines de transformation de l'asclépiade était à Granby. Elle a fermé avec "
        "le reste de la filière, et l'Estrie compte encore des producteurs de la plante.",
    "Montérégie":
        "La Montérégie est une des régions où l'asclépiade se cultive encore, malgré "
        "l'effondrement de la filière en 2018.",
    "Saguenay - Lac-Saint-Jean":
        "Un de nos fournisseurs d'asclépiade cultive au Lac-Saint-Jean depuis nos débuts, et "
        "il l'est encore aujourd'hui. Pour une filière qui s'est effondrée en 2018, ce n'est "
        "pas banal.",
    "Capitale-Nationale/Chaudière-Appalaches":
        "Notre atelier est dans Limoilou, et c'est là que la soie d'asclépiade devient de "
        "l'isolant. C'est la seule étape qu'on a refusé de délocaliser.",
    "Est-du-Québec":
        "L'asclépiade est une vivace indigène qui pousse avec peu d'intrants, ce qui la rend "
        "intéressante pour des régions où les cultures conventionnelles rendent mal.",
}

# --- accroches thematiques, quand la region ne mord pas ---------------------

THEMES = {
    "C": "En 2018, la filière québécoise de l'asclépiade s'est effondrée et des producteurs "
         "ont retourné leurs champs. Huit ans plus tard, on achète encore de l'asclépiade "
         "québécoise et on la transforme nous-mêmes à Québec.",
    "D": "Le monarque est en voie de disparition au Canada depuis un statut confirmé en 2024, "
         "et son déclin est largement attribué à la disparition de l'asclépiade dans les "
         "zones agricoles.",
    "E": "En 2021, on a acheté une matelasseuse usagée en Ontario et bâti nos propres procédés "
         "parce qu'aucun sous-traitant ne voulait toucher à l'asclépiade. En 2025, on a "
         "confié l'assemblage textile à des manufacturiers externes.",
    "F": "La soie d'asclépiade est une fibre creuse et hydrophobe attachée aux graines de la "
         "plante. À poids égal, la littérature la donne pour environ 10 % plus isolante que "
         "le duvet. C'est un repère de laboratoire, pas une promesse de manteau.",
    "G": "L'asclépiade est la plante que les agriculteurs arrachent depuis 50 ans. La soie "
         "attachée à ses graines est ce qui isole nos manteaux, nos mitaines et nos tuques.",
    "I": "Lasclay est une entreprise de Québec qui transforme la soie d'asclépiade en isolant "
         "textile.",
    "B": "Lasclay est une entreprise de Québec qui transforme la soie d'asclépiade en isolant "
         "textile.",
}

# --- corps par angle --------------------------------------------------------

CORPS = {
    "B": "Je suis disponible pour une entrevue avant ou après la diffusion, et je peux parler "
         "de ce qui se fait encore ici comme de ce qui n'y est plus.",
    "C": "Ce n'est pas la télévision qui devrait intéresser vos lecteurs. Le sujet, c'est "
         "plutôt : est-ce qu'il y a enfin un acheteur stable au bout du champ, et à quelles "
         "conditions. Je peux parler des volumes, de la fenêtre de récolte et de ce qui bloque "
         "encore côté mécanisation, sans rien promettre que je ne peux pas tenir.",
    "D": "Le pari de Lasclay est économique avant d'être écologique : si l'asclépiade paie, les "
         "agriculteurs la gardent dans leurs champs. On a distribué environ 10 millions de "
         "graines en Amérique du Nord.\n\nJe précise toujours la même chose, parce que c'est "
         "souvent mal raconté : acheter un produit ne sauve pas un papillon en particulier. Le "
         "lien est systémique.",
    "E": "C'est ce qui a permis un manteau isolé à l'asclépiade autour de 300 $, après des "
         "tentatives de l'industrie à plus de 500 $ et 1000 $ qui ont toutes été "
         "discontinuées.\n\nCe dont je peux parler : industrialiser une matière qui n'a pas de "
         "chaîne d'approvisionnement, et ce que le virage a coûté.",
    "F": "C'est pour ça que je préfère que vous testiez plutôt que de me croire. Je peux vous "
         "envoyer un produit avant l'hiver, sans condition et sans droit de regard sur ce que "
         "vous en écrirez.",
    "G": "J'ai du matériel photo qui montre bien le contraste : la gousse dans le champ, la "
         "soie dans la main, le produit porté. L'atelier de Limoilou est ouvert si vous voulez "
         "voir la matière.\n\nUne précision qui compte : l'isolant est cultivé et transformé au "
         "Québec, l'assemblage des produits finis se fait en Tunisie depuis 2025. Je ne dis pas "
         "« fabriqué au Québec », ce serait faux.",
    "I": "Je suis disponible pour une entrevue la semaine du 14 septembre et le vendredi 18 au "
         "matin.\n\nLe sujet en une phrase : une entreprise de Québec va expliquer à un "
         "auditoire pancanadien pourquoi la mauvaise herbe des champs de maïs peut isoler un "
         "manteau.",
    "H": "I cannot say how the pitch went before it airs. I can talk about everything else: why "
         "Quebec's first milkweed industry collapsed in 2018, and what it takes to build a "
         "supply chain for a material that does not have one.",
}

CONFIDENTIEL = "Je ne peux pas dire comment ça s'est terminé avant la diffusion."

# --- textes ecrits a la main ------------------------------------------------
# Contacts a forte valeur, ou l'assemblage ne suffirait pas.

ECRITS_A_LA_MAIN = {
"jean-michel_leprince@radio-canada.ca": """Bonjour Jean-Michel,

Vous couvrez l'asclépiade depuis au moins 2014. Le Téléjournal, la soie d'Amérique
partie sur l'Everest, le pouvoir absorbant de la fibre sur les hydrocarbures, le lien
avec le monarque. Vous avez suivi cette histoire plus longtemps que la plupart des
entreprises qui s'y sont essayées.

Vous avez donc vu la suite : Encore 3, Fibre Monark et Protec-Style ont fermé les unes
après les autres, et la coopérative Monark s'est retrouvée sans acheteur.

On est l'entreprise qui a démarré après. Six ans plus tard, on achète encore de
l'asclépiade québécoise, on transforme l'isolant nous-mêmes à Limoilou, et {DIFF}.

{CONF}

Ce que je peux raconter, et que personne n'a encore raconté au complet : pourquoi la
première filière est morte, ce qu'on a fait différemment, et ce qui n'est toujours pas
réglé, à commencer par la mécanisation de la récolte. Je n'ai pas besoin de vous
expliquer la plante.""",

"redaction.mekinac@lebulletindeschenaux.com": """Bonjour Réjean,

Vous couvrez Mékinac. C'est là que l'histoire industrielle de l'asclépiade au Québec
s'est jouée, et c'est là qu'elle s'est cassée : Les Industries Encore 3, à Saint-Tite,
achetaient 90 % des récoltes du Québec avant de faire faillite. Des producteurs de
votre secteur se sont retrouvés avec des champs et aucun acheteur.

On est l'entreprise qui a démarré après cet effondrement. Six ans plus tard, on achète
encore de l'asclépiade québécoise et on la transforme nous-mêmes à Québec.

{DIFF}. {CONF}

Ce qui devrait intéresser vos lecteurs, ce n'est pas la télévision. C'est de savoir
s'il y a de nouveau un débouché pour la plante qui a fait tant de promesses dans
Mékinac. Ma réponse honnête : il y en a un, il est plus petit que ce qu'on avait promis
en 2014, et je ne peux garantir aucun volume à long terme. Je peux en parler
franchement.""",

"gabriel.delisle@lenouvelliste.qc.ca": """Bonjour Gabriel,

L'asclépiade a été une histoire mauricienne avant d'être la nôtre. L'usine de
Saint-Tite achetait 90 % des récoltes du Québec, puis Les Industries Encore 3 ont fait
faillite et la filière s'est effondrée.

On a démarré après. Aujourd'hui on achète encore de l'asclépiade québécoise, on
transforme l'isolant à Limoilou, et {DIFF}.

{CONF}

L'angle qui me semble le plus juste pour la Mauricie : huit ans après, est-ce qu'il
reste quelque chose de la promesse de Saint-Tite. Je peux répondre sans enjoliver, y
compris sur ce qui n'est toujours pas réglé.""",

"jlafrance@lenouvelliste.qc.ca": """Bonjour Jacinthe,

Vous êtes en Mauricie, la région où l'asclépiade a eu son grand moment industriel avant
de tout perdre : l'usine de Saint-Tite achetait 90 % des récoltes du Québec, puis elle
a fermé.

On est l'entreprise qui a démarré après. On transforme la soie en isolant à Québec et
on en fait des manteaux, des mitaines, des sacs isothermes. {DIFF}.

{CONF}

Côté consommation, la question que je trouve la plus honnête : est-ce qu'un produit en
asclépiade vaut son prix. Je peux y répondre en expliquant ce que la fibre fait
réellement, et ce qu'elle ne fait pas.""",

"paule.vermot-desroches@lenouvelliste.qc.ca": """Bonjour Paule,

L'asclépiade est un dossier économique mauricien avant tout : Les Industries Encore 3,
à Saint-Tite, achetaient 90 % des récoltes du Québec quand elles ont fait faillite,
entraînant la filière avec elles.

On a démarré après cet effondrement, avec l'idée de commencer petit et de vendre des
produits finis avant de promettre une industrie. {DIFF}.

{CONF}

Le sujet économique : à quoi ressemble l'industrialisation d'un matériau qui n'a aucune
chaîne d'approvisionnement. On a rapatrié notre production en 2021 faute de
sous-traitant, puis sorti l'assemblage textile du Québec en 2025 pour rendre un manteau
accessible à 300 $. Les deux décisions se défendent, et elles se contredisent en
apparence.""",

"cgermain@ledevoir.com": """Bonjour Chloé,

Vous êtes en Mauricie, où l'asclépiade a eu son usine et sa faillite : Saint-Tite
achetait 90 % des récoltes du Québec avant que Les Industries Encore 3 ferment.

On est l'entreprise qui a démarré après. On achète encore de l'asclépiade québécoise,
on transforme l'isolant à Limoilou, et {DIFF}.

{CONF}

Il y a un sujet agricole réel là-dedans : les producteurs qui ont survécu à 2018 sont
encore là, et la question qu'ils posent n'a pas changé. Est-ce qu'il y a un acheteur
stable au bout du champ. Je peux répondre sans rien promettre que je ne peux tenir.""",

"karine.tremblay@latribune.qc.ca": """Bonjour Karine,

Une des usines de transformation de l'asclépiade était à Granby. Elle a fermé avec le
reste de la filière en 2018, et l'Estrie compte encore des producteurs de la plante.

On est l'entreprise qui a démarré après cet effondrement. La Tribune a d'ailleurs repris
en décembre dernier un article sur notre dilemme de fabrication.

{DIFF}. {CONF}

Ce que je peux raconter : pourquoi la première filière est morte, ce qu'on a fait
différemment, et ce que ça coûte de rendre un manteau d'asclépiade accessible à 300 $.""",

"marie.allard@laesdebrouillards.com": """Bonjour Marie,

Il y a une histoire de sciences dans l'asclépiade qui se raconte bien à tous les âges :
la soie attachée à ses graines n'est pas un poil, c'est un tube creux enduit d'une cire
hydrophobe, et les fibres portent une charge qui les fait se repousser. C'est ce qui
forme le parachute autour de la graine, et c'est ce qui emprisonne l'air.

Chaque follicule produit en moyenne plus de 200 graines soyeuses. C'est aussi la seule
plante que les chenilles du monarque peuvent manger.

On en fait de l'isolant pour des manteaux et des mitaines, et {DIFF}.

{CONF} Si le sujet vous intéresse pour vos lecteurs, l'atelier de Limoilou est ouvert.""",
}

# Contacts sans courriel : le brouillon existe, l'adresse manque.
SANS_COURRIEL_MANUEL = {
"Antoine Stab": """Bonjour Antoine,

Vous aviez écrit sur le soyer du Québec dans Espaces en février 2015, à l'époque où la
fibre devait remplacer le duvet. Vous savez donc ce qui a suivi : les transformateurs
ont fait faillite les uns après les autres à partir de 2017.

On est l'entreprise qui a démarré après. Six ans plus tard, on achète encore de
l'asclépiade québécoise, on transforme l'isolant nous-mêmes, et {DIFF}.

{CONF}

Côté plein air, ce qui a changé depuis votre article : il existe maintenant un manteau
à inserts d'asclépiade amovibles autour de 300 $. Je préfère que vous le testiez plutôt
que de me croire.""",
}


def accroche(region, angle):
    """Premiere ligne. La region gagne quand elle a une histoire reelle."""
    if region in REGIONS and angle != "I":
        return REGIONS[region]
    return THEMES.get(angle, THEMES["I"])


def brouillon(prenom, angle, region, courriel, nom):
    manuel = ECRITS_A_LA_MAIN.get((courriel or "").strip().lower())
    if not manuel and not courriel:
        manuel = SANS_COURRIEL_MANUEL.get(f"{prenom} {nom}".strip())
    if manuel:
        return manuel.format(DIFF=DIFFUSION, CONF=CONFIDENTIEL)

    salut = f"Bonjour {prenom}," if prenom else "Bonjour,"
    tete = accroche(region, angle)
    diff = DIFFUSION[0].upper() + DIFFUSION[1:]
    corps = CORPS.get(angle, CORPS["I"])

    if angle in ("B", "I"):
        return f"{salut}\n\n{tete} {diff}.\n\n{CONFIDENTIEL}\n\n{corps}"
    return f"{salut}\n\n{tete}\n\n{diff}.\n\n{CONFIDENTIEL}\n\n{corps}"


def main(src, dst_xlsx, dst_md):
    wb = openpyxl.load_workbook(src)
    wf = next(w for w in wb.worksheets if w.title.startswith("Liste froide"))
    col = wf.max_column + 1
    c = wf.cell(1, col, "Brouillon")
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor="E3EADF")
    wf.column_dimensions[c.column_letter].width = 70

    par_angle, mains, sans_courriel = {}, 0, 0
    lignes_md = ["# Brouillons, liste froide FPJQ",
                 "",
                 "Un brouillon par contact. Les textes marqués **écrit à la main** valent la peine "
                 "d'être relus avant l'envoi : ce sont les contacts qui ont couvert l'asclépiade ou "
                 "qui sont géographiquement au cœur de l'effondrement de la filière.",
                 "",
                 "Les autres viennent d'un assemblage : une accroche régionale quand la région a une "
                 "histoire réelle avec l'asclépiade, une accroche thématique sinon, puis le corps de "
                 "l'angle assigné. Chaque première ligne diffère vraiment, comme l'exige le rythme "
                 "d'envoi.",
                 ""]

    for i, row in enumerate(wf.iter_rows(min_row=2, values_only=True), start=2):
        if not row[2]:
            continue
        prio, prenom, nom, fonction, media, courriel = row[0:6]
        region, angle = row[7], row[14]
        if angle in (None, "—"):
            continue
        txt = brouillon(prenom, angle, region, courriel, nom)
        wf.cell(i, col, txt).alignment = Alignment(wrap_text=True, vertical="top")
        par_angle.setdefault(angle, []).append(
            (prio, prenom, nom, media, courriel, region, txt))
        est_manuel = (courriel or "").strip().lower() in ECRITS_A_LA_MAIN \
            or f"{prenom} {nom}".strip() in SANS_COURRIEL_MANUEL
        mains += est_manuel
        sans_courriel += not courriel

    for angle in sorted(par_angle):
        contacts = par_angle[angle]
        lignes_md += [f"## Angle {angle} — {len(contacts)} contacts", ""]
        for prio, prenom, nom, media, courriel, region, txt in contacts:
            est_manuel = (courriel or "").strip().lower() in ECRITS_A_LA_MAIN \
                or f"{prenom} {nom}".strip() in SANS_COURRIEL_MANUEL
            tag = " — **écrit à la main**" if est_manuel else ""
            adr = f"`{courriel}`" if courriel else "*adresse à trouver*"
            lignes_md += [f"### {prenom} {nom} — {media} — {adr}{tag}",
                          f"*Priorité {prio} · {region}*", "", "```", txt, "```", ""]

    wb.save(dst_xlsx)
    with open(dst_md, "w") as f:
        f.write("\n".join(lignes_md))

    total = sum(len(v) for v in par_angle.values())
    print(f"{total} brouillons écrits dans {dst_xlsx} et {dst_md}")
    print(f"  dont {mains} écrits à la main")
    print(f"  dont {sans_courriel} sans adresse courriel")
    for a in sorted(par_angle):
        print(f"  angle {a} : {len(par_angle[a])}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3])

#!/usr/bin/env python3
"""Aplatit les deux feuilles de Lasclay_v2.xlsx en un seul chiffrier a editer.

Gabriel edite la colonne « TA VERSION » pour montrer le registre attendu ; tout
ce qu'il a deja ecrit dans les trois colonnes de droite est repris d'une version
a l'autre, sinon chaque regeneration effacerait ses corrections.
"""
import sys
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

VERSION = 12
SOURCE = "Lasclay_v2.xlsx"
CIBLE = "Lasclay_brouillons_a_editer.xlsx"

COLONNES = ["Liste", "Nom", "Média", "Courriel", "Région", "Date / fonction",
            "Contexte connu", "Angle", "Registre", "Objet", f"Brouillon v{VERSION}",
            "TA VERSION", "CE QUI N'ALLAIT PAS", "Verdict"]
LARGEURS = [11, 24, 24, 32, 20, 15, 30, 7, 10, 46, 86, 86, 40, 14]

def changements(kit, chauds, mercis):
  return [
    (f"Version {VERSION} — tes dix corrections", None),
    (None, None),
    ("Le média kit",
     "« Notre média kit est ici » annonçait un envoi que personne n'avait demandé. Ta formule "
     "remet la décision au journaliste et ouvre la porte au collègue qui prendra peut-être le "
     f"sujet : « Si ça vous intéresse de couvrir (ou un.e collègue?)… ». Dans les {kit} "
     "brouillons."),
    ("L'ordre",
     "Plus une seule allusion à la diffusion avant qu'elle soit annoncée. « Cette année, la "
     "date tombe presque au même endroit » et « la diffusion tombe cinq ans jour pour jour "
     "après votre article » sautent : une coïncidence de calendrier ne passe pas devant la "
     "nouvelle."),
    ("Les remerciements",
     f"Les {mercis} remerciements sont raccourcis et collés au paragraphe du rappel. Plus de "
     "compliment sur un choix éditorial — « merci d'avoir ouvert le cahier climatique à une "
     "plante » se lit comme une lecture de dossier. Un fait ne reste que s'il dit ce que "
     "l'article a donné : « le texte a beaucoup circulé chez nos clients »."),
    ("Les missions",
     "« Depuis, tout mon temps va à mes 2 grandes missions » parlait de ton emploi du temps. "
     "« On se concentre sur nos 2 grandes missions » parle de l'entreprise, comme dans ta "
     "correction à Fanny Samson."),
    ("Tes dix textes",
     "Larocque, Bérubé, Pouliot, Paquet, Anne-Sophie Roy, Lafrance, Marie Tison, Simard, "
     "Samson et Lemieux sont les tiens, mot pour mot. Seule la ligne du média kit y est "
     "uniformisée, là où tu ne l'avais pas encore changée."),
    ("Registre",
     "Melissa Paquet, Anne-Sophie Roy et Annie Lafrance passent au « tu », comme tes textes. "
     "Avec Bérubé et Pouliot, ça fait cinq."),
    ("Sorties",
     "Sylvie Lacombe, Justine Friis et Francis Higgins ne sont plus dans ton chiffrier : je "
     "les ai retirés. Dis-le si c'était un accident et je les remets."),
    ("Jean-Michel Leprince",
     "Promu de la liste froide à la liste chaude. Son adresse reste à confirmer : elle vient "
     "d'une recherche web, pas d'un échange."),
    (None, None),
    ("Ce qui reste à vérifier", None),
    (None, None),
    ("Ne pas envoyer",
     "Anne-Sophie Roy à l'adresse Québecor, doublon de sa ligne Radio-Canada."),
    ("Six salutations", "Fiches FPJQ mal formées, signalées dans la colonne Précaution."),
    ("Faits nouveaux",
     "Tes corrections ajoutent deux choses que je n'avais pas : les soins pour la peau à "
     "l'huile de graines d'asclépiade, et « mitaines » plutôt que « moufles ». Les deux sont "
     "corrigés là où tu les as écrits, mais pas ailleurs — dis-moi si le catalogue doit être "
     "revu partout."),
    (None, None),
    ("Rappels", None),
    (None, None),
    ("Diffusion", "Jeudi 17 septembre 2026, 20 h (20 h 30 NT), CBC et CBC Gem. "
                  "Prévente le samedi 12 septembre à 9 h."),
    ("Envoi", "Missive depuis media@lasclay.com, un appel par journaliste, suivi des "
              "ouvertures et des clics désactivé, environ 60 par jour. Le mode par "
              "défaut dépose un brouillon."),
    ("Interdits CBC", "Rien sur l'issue avant le 17. Aucun logo ni « vu à Dragons' Den »."),
]


def index(ws):
    return {c.value: i for i, c in enumerate(ws[1])}


def lignes():
    wb = openpyxl.load_workbook(SOURCE)
    ws = wb["Liste chaude (34)"]
    h = index(ws)
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[h["Nom"]]:
            continue
        yield ["chaude", r[h["Nom"]], r[h["Média"]], r[h["Courriel"]], None,
               r[h["Dernier contact"]], r[h["Sujet du dernier échange"]],
               r[h["Angle"]], r[h["Registre"]], r[h["Objet suggéré"]], r[h["Brouillon"]]]

    ws = wb["Liste froide FPJQ (217)"]
    h = index(ws)
    for r in ws.iter_rows(min_row=2, values_only=True):
        nom = " ".join(x for x in (r[h["Prénom"]], r[h["Nom"]]) if x)
        if not nom:
            continue
        yield [f"froide {r[h['Priorité']]}", nom, r[h["Média"]], r[h["Courriel"]],
               r[h["Région"]], r[h["Fonction"]], r[h["Secteurs pertinents"]],
               r[h["Angle"]], "vous", r[h["Objet suggéré"]], r[h["Brouillon"]]]


def acquis():
    """Ce que Gabriel a deja ecrit dans les trois colonnes de droite, par courriel."""
    try:
        wb = openpyxl.load_workbook(CIBLE)
    except FileNotFoundError:
        return {}
    ws = wb["Brouillons à éditer"]
    garde = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[3] and any(r[11:14]):
            garde[r[3]] = list(r[11:14])
    return garde


def main():
    garde = acquis()
    # Les comptes du sommaire se calculent, ils ne s'écrivent pas : une sortie de
    # plus les laissait faux, et un chiffre faux dans un sommaire est pire que pas
    # de sommaire du tout.
    tout = list(lignes())
    kit = sum(1 for l in tout if l[10] and "1pyCUbfHYQhpXXl4FoCC2RCFXKRvGS5Zr" in l[10])
    chauds = sum(1 for l in tout if l[0] == "chaude" and not l[10].startswith("NE PAS"))
    mercis = sum(1 for l in tout if l[0] == "chaude" and "erci" in l[10]
                 and not l[10].startswith("NE PAS"))
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Ce qui a changé"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 108
    for titre, texte in changements(kit, chauds, mercis):
        ws.append([titre, texte])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=12 if titre and not texte else 11)
        ws.cell(ws.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("Brouillons à éditer")
    ws.append(COLONNES)
    tete = PatternFill("solid", fgColor="1F3864")
    for i, c in enumerate(ws[1], start=1):
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = tete
        c.alignment = Alignment(vertical="center")
        ws.column_dimensions[c.column_letter].width = LARGEURS[i - 1]

    edite = PatternFill("solid", fgColor="FFF2CC")
    n = 0
    for ligne in lignes():
        ws.append(ligne + garde.get(ligne[3], [None, None, None]))
        n += 1
        for col in (7, 10, 11, 12, 13):
            ws.cell(ws.max_row, col).alignment = Alignment(wrap_text=True, vertical="top")
        for col in (12, 13, 14):
            ws.cell(ws.max_row, col).fill = edite

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:N{ws.max_row}"
    verdict = DataValidation(type="list",
                             formula1='"à réécrire au complet,à corriger,bon,ne pas envoyer"',
                             allow_blank=True)
    registre = DataValidation(type="list", formula1='"tu,vous,—"', allow_blank=True)
    ws.add_data_validation(verdict)
    ws.add_data_validation(registre)
    verdict.add(f"N2:N{ws.max_row}")
    registre.add(f"I2:I{ws.max_row}")

    wb.save(CIBLE)
    print(f"{CIBLE} : {n} contacts, {kit} brouillons avec le média kit, "
          f"{len(garde)} édition(s) reprise(s)")


if __name__ == "__main__":
    sys.exit(main())

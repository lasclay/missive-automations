#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
verifier-liens-ahrefs.py

À exécuter dès que la politique réseau autorise api.ahrefs.com.
Fait en une seule passe ce que 72 ouvertures de page feraient mal :

  1. tire les domaines référents de lasclay.com depuis l'API Ahrefs ;
  2. les croise avec la colonne « URL de la page cible » du chiffrier et
     remplit la colonne « Lien en place ? » ;
  3. optionnel : sort l'écart de domaines référents entre lasclay.com et des
     domaines comparables, ce qui produit des occasions prouvées plutôt que
     des hypothèses.

Usage :
    export AHREFS_TOKEN="..."
    python3 verifier-liens-ahrefs.py chiffrier.xlsx
    python3 verifier-liens-ahrefs.py chiffrier.xlsx --gap coopmonark.com autredomaine.com

Le script n'écrase jamais le fichier d'entrée : il écrit une copie datée et un
journal avant/après en JSON, conformément à la règle du journal de la zone verte.

Avant la première exécution, vérifier deux choses dans la documentation Ahrefs :
le chemin exact du point d'accès et les noms de champs de `select`. Ils ont
changé entre les versions de l'API et le script ne peut pas le deviner.
"""

import json
import os
import sys
import time
from datetime import date
from urllib.parse import urlparse

import requests
import openpyxl

API = "https://api.ahrefs.com/v3"
TOKEN = os.environ.get("AHREFS_TOKEN")
COL_URL = 5        # E
COL_ETAT = 20      # T, « Lien en place ? »
PREMIERE, DERNIERE = 5, 76


def entetes():
    if not TOKEN:
        sys.exit("AHREFS_TOKEN absent de l'environnement. Rien n'a été fait.")
    return {"Authorization": "Bearer " + TOKEN, "Accept": "application/json"}


def domaine(url):
    """Domaine nu, sans www, à partir d'une URL de la colonne E."""
    if not url:
        return None
    u = str(url).strip()
    if not u.startswith("http"):
        u = "https://" + u
    h = urlparse(u).netloc.lower()
    return h[4:] if h.startswith("www.") else h


def refdomains(cible, limite=1000):
    """Domaines qui pointent vers `cible`. Pagination simple."""
    resultats, offset = {}, 0
    while True:
        r = requests.get(
            API + "/site-explorer/refdomains",
            headers=entetes(),
            params={
                "target": cible,
                "mode": "domain",
                "protocol": "both",
                "limit": min(limite, 1000),
                "offset": offset,
                "select": "domain,domain_rating,links_to_target,first_seen",
            },
            timeout=60,
        )
        if r.status_code == 429:
            time.sleep(10)
            continue
        r.raise_for_status()
        lot = r.json().get("refdomains", [])
        if not lot:
            break
        for d in lot:
            resultats[(d.get("domain") or "").lower()] = d
        if len(lot) < 1000:
            break
        offset += len(lot)
    return resultats


def verifier(chemin, cible="lasclay.com"):
    refs = refdomains(cible)
    print("domaines référents trouvés pour %s : %d" % (cible, len(refs)))

    wb = openpyxl.load_workbook(chemin)
    ws = wb["Opportunités"]
    journal = {"date": str(date.today()), "cible": cible,
               "domaines_referents": len(refs), "lignes": []}

    for r in range(PREMIERE, DERNIERE + 1):
        if ws.cell(r, 4).value is None:
            continue
        d = domaine(ws.cell(r, COL_URL).value)
        avant = ws.cell(r, COL_ETAT).value
        if not d:
            apres = "Pas d'URL à vérifier"
        elif d in refs:
            info = refs[d]
            apres = "Oui, lien actif (DR %s, %s lien(s))" % (
                info.get("domain_rating", "?"), info.get("links_to_target", "?"))
        else:
            apres = "Non, aucun lien depuis ce domaine"
        ws.cell(r, COL_ETAT, apres)
        journal["lignes"].append(
            {"ligne": r, "organisation": ws.cell(r, 4).value,
             "domaine": d, "avant": avant, "apres": apres})

    sortie = chemin.replace(".xlsx", "-verifie-%s.xlsx" % date.today())
    wb.save(sortie)
    with open(sortie.replace(".xlsx", "-journal.json"), "w", encoding="utf-8") as f:
        json.dump(journal, f, ensure_ascii=False, indent=2)

    deja = sum(1 for l in journal["lignes"] if l["apres"].startswith("Oui"))
    print("lien déjà en place : %d lignes" % deja)
    print("à démarcher : %d lignes" % (len(journal["lignes"]) - deja))
    print("écrit : " + sortie)
    return refs


def ecart(refs_lasclay, comparables):
    """Domaines qui pointent vers un comparable mais pas vers lasclay.com.

    C'est la prospection la plus rentable : chaque domaine sorti d'ici a déjà
    prouvé qu'il accepte de lier un acteur de ce secteur.
    """
    trouve = {}
    for c in comparables:
        print("écart avec %s..." % c)
        for d, info in refdomains(c).items():
            if d and d not in refs_lasclay:
                trouve.setdefault(d, {"domaine": d, "dr": info.get("domain_rating"),
                                      "vu_chez": []})["vu_chez"].append(c)
    classe = sorted(trouve.values(),
                    key=lambda x: (-len(x["vu_chez"]), -(x["dr"] or 0)))
    nom = "ecart-domaines-%s.json" % date.today()
    with open(nom, "w", encoding="utf-8") as f:
        json.dump(classe, f, ensure_ascii=False, indent=2)
    print("%d domaines à évaluer, écrits dans %s" % (len(classe), nom))
    print("Les domaines vus chez plusieurs comparables sont les plus prometteurs.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    fichier = sys.argv[1]
    refs = verifier(fichier)
    if "--gap" in sys.argv:
        comparables = sys.argv[sys.argv.index("--gap") + 1:]
        if comparables:
            ecart(refs, comparables)

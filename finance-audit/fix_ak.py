#!/usr/bin/env python3
"""Phase AK — le bloc 2028-2029 manquant, et les blocs périmés masqués.

« Ventes prévisionnelles » porte un bloc de quinze colonnes par exercice :
croissance, quantité annuelle, douze mois, total. Il y en avait pour 2023-2024,
2024-2025, 2025-2026, 2026-2027 et 2027-2028 — mais pas pour 2028-2029, alors que
le résultat va jusqu'en août 2029. FY2029 se calculait donc à part, en multipliant
le profil mensuel de FY2028 par un facteur unique, ce qui rendait la dernière
année du plan invisible dans la feuille où on va la chercher.

Le bloc 2028-2029 est construit en répliquant celui de 2027-2028 décalé de quinze
colonnes — exactement la relation qu'entretiennent déjà les blocs entre eux. Les
six facteurs de croissance par catégorie pointent tous sur Inputs!C54, le facteur
FY2029 : le résultat est identique à ce que le modèle calculait, mais il devient
lisible et modifiable produit par produit.

Les deux blocs périmés (2023-2024 et 2024-2025, dont les totaux sont à zéro) et le
bloc 2025-2026, désormais piloté par le réel de QuickBooks, sont masqués.
"""
import re
import sys
sys.path.insert(0, 'tools')
from openpyxl.utils import column_index_from_string, get_column_letter
from xledit import Editor
import xlcalc

V = 'Ventes prévisionnelles '
PNL = 'Résultats-Prev 2025-2029'
F = 'PREVISIONS LASCLAY - version audit 2026-07-30.xlsx'

SHIFT = 15                       # largeur d'un bloc d'exercice
SRC = ['BS', 'BT'] + ['BU', 'BV', 'BW', 'BX', 'BY', 'BZ', 'CA', 'CB', 'CC', 'CD',
                      'CE', 'CF'] + ['CG']
DST = [get_column_letter(column_index_from_string(c) + SHIFT) for c in SRC]
MOIS29 = ['2028-09', '2028-10', '2028-11', 'Dec. 2028', 'Jan. 2029', 'Feb. 2029',
          'Mar. 2029', 'Apr. 2029', 'May 2029', 'Jun. 2029', 'Jul. 2029', 'Aug. 2029']
CROISS = [10, 83, 141, 155, 186, 243]      # rangées des facteurs de croissance
# blocs à masquer : (première colonne, dernière colonne)
MASQUER = [('A', 'R'), ('T', 'AK'), ('AM', 'BD')]
REF = re.compile(r'(?<![A-Z0-9_!])(\$?)([A-Z]{1,3})(\$?)(\d+)(?![(\d])')


def shift_formula(f, delta=SHIFT):
    """Décale de `delta` colonnes toutes les références À CETTE FEUILLE.

    Les références absolues ($BT$8) se décalent aussi : dans un bloc répliqué,
    elles pointent sur la colonne d'ancrage du bloc, qui bouge avec lui. Les
    références à une autre feuille (Inputs!C54) ne bougent pas."""
    out, last = [], 0
    for m in re.finditer(r"'[^']*'!\$?[A-Z]{1,3}\$?\d+|[A-Za-z_]\w*!\$?[A-Z]{1,3}\$?\d+", f):
        out.append(shift_local(f[last:m.start()], delta))
        out.append(m.group(0))               # référence externe : intacte
        last = m.end()
    out.append(shift_local(f[last:], delta))
    return ''.join(out)


def shift_local(part, delta):
    def rep(m):
        d1, col, d2, row = m.groups()
        i = column_index_from_string(col) + delta
        return f'{d1}{get_column_letter(i)}{d2}{row}'
    return REF.sub(rep, part)


def build(dst=F, src=F):
    bk = xlcalc.load(src)
    e = Editor(src)
    e.expand_shared()

    # --- AK1 : répliquer le bloc 2027-2028 en 2028-2029 -------------------
    n = 0
    for r in range(1, 660):
        for cs, cd in zip(SRC, DST):
            a, b = f'{cs}{r}', f'{cd}{r}'
            f = bk.formulas[V].get(a)
            v = bk.values[V].get(a)
            if f:
                e.set(V, b, '=' + shift_formula(f))
                n += 1
            elif v is not None and not isinstance(v, str):
                e.set(V, b, float(v))
                n += 1
            elif isinstance(v, str):
                e.set(V, b, v)
                n += 1

    # --- AK2 : les en-têtes du nouveau bloc -------------------------------
    for c, lab in zip(DST[2:14], MOIS29):
        e.set(V, f'{c}1', lab)
        e.set(V, f'{c}2', 'Forecast')
    e.set(V, f'{DST[14]}1', 'Total 2028-2029')
    e.set(V, f'{DST[14]}2', 'Forecast')
    e.set(V, f'{DST[0]}1', None)
    e.set(V, f'{DST[1]}1', None)

    # --- AK3 : un seul facteur de croissance pour FY2029 ------------------
    # Inputs!C54 bascule déjà entre les deux scénarios ; les six catégories le
    # partagent, comme le faisait l'ancien calcul.
    for r in CROISS:
        e.set(V, f'{DST[0]}{r}', '=Inputs!$C$54')

    # --- AK4 : le résultat lit le nouveau bloc ----------------------------
    tot = DST[14]
    parts = {}
    for i, c in enumerate(['AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN',
                           'AO', 'AP', 'AQ']):
        f = bk.formulas[PNL].get(f'{c}5') or ''
        m = re.search(r'\*([0-9.]+)$', f)
        parts[i] = m.group(1) if m else None
    FY29 = ['AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE']
    for i, c in enumerate(FY29):
        if parts[i]:
            e.set(PNL, f'{c}5', f"='{V}'!${tot}$5*{parts[i]}")
    e.set(PNL, 'C5', 'Budget de FY2026 tel qu’il avait été posé (constantes) : un '
                     'exercice fermé ne se rebudgète pas. FY2027 à FY2029 sont pilotés '
                     'par « Ventes prévisionnelles », un bloc de 15 colonnes par exercice.')

    # --- AK5 : masquer les blocs périmés et l'exercice fermé --------------
    xml = e.parts[e.sheetpart[V]].decode('utf8')
    cols = []
    for c1, c2 in MASQUER:
        cols.append((column_index_from_string(c1), column_index_from_string(c2)))
    tag = ''.join(f'<col min="{a}" max="{b}" hidden="1" customWidth="1" width="8"/>'
                  for a, b in cols)
    if '<cols>' in xml:
        xml = xml.replace('<cols>', '<cols>' + tag, 1)
    else:
        i = xml.index('<sheetData>')
        xml = xml[:i] + '<cols>' + tag + '</cols>' + xml[i:]
    e.parts[e.sheetpart[V]] = xml.encode('utf8')

    e.set_full_calc()
    e.save(dst)
    print(f'{n} cellules répliquées dans le bloc 2028-2029')
    return e


if __name__ == '__main__':
    build()
    bk = xlcalc.load(F)
    tot = DST[14]
    print(f'{"bloc":<12}{"total des ventes":>18}')
    for c, lab in (('Q', '2023-2024'), ('AJ', '2024-2025'), ('BC', '2025-2026'),
                   ('BR', '2026-2027'), ('CG', '2027-2028'), (tot, '2028-2029')):
        print(f'{lab:<12}{bk.get(V, c + "5"):>18,.0f}')
    print()
    for lbl, t in (('FY2026', 'P'), ('FY2027', 'AD'), ('FY2028', 'AR'), ('FY2029', 'BF')):
        print(f'{lbl}  revenus prévus {bk.get(PNL, t + "5"):>12,.0f}   '
              f'ventes nettes {bk.get(PNL, t + "26"):>12,.0f}')

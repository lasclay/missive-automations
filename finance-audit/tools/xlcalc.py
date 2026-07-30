#!/usr/bin/env python3
"""Minimal spreadsheet formula evaluator for the Lasclay forecast workbook.

LibreOffice cannot load this file (it stalls before calculating), so the audit
needs its own engine. The workbook's vocabulary is small: SUMIF, SUM, IFERROR,
IF, AND, OR, MAX, MIN, MONTH, SEARCH, ISNUMBER, TRUE plus arithmetic. Everything
here is lazy + memoized, with cycle detection.
"""
import re, datetime, math
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

ERRORS = {'#REF!', '#DIV/0!', '#VALUE!', '#NAME?', '#N/A', '#NULL!', '#NUM!'}


class Err(str):
    __slots__ = ()


class Blank(float):
    """An empty cell. Numifies to 0 but equals "" in a comparison, the way
    Excel treats it — the workbook's staffing windows rely on OR($E$46="",...)."""
    __slots__ = ()


BLANK = Blank(0.0)


# ---------------------------------------------------------------- tokenizer
TOKEN_RE = re.compile(r"""
    (?P<str>"(?:[^"]|"")*")
  | (?P<err>\#(?:REF!|DIV/0!|VALUE!|NAME\?|N/A|NULL!|NUM!))
  | (?P<sheet>'(?:[^']|'')*'!|[A-Za-z_][A-Za-z0-9_.]*!)
  | (?P<num>\d+(?:\.\d+)?(?:[eE][-+]?\d+)?|\.\d+)
  | (?P<ref>\$?[A-Z]{1,3}\$?\d{1,7})
  | (?P<colrange>\$?[A-Z]{1,3}:\$?[A-Z]{1,3})
  | (?P<func>[A-Za-z_][A-Za-z0-9_.]*\()
  | (?P<op><=|>=|<>|[-+*/^&<>=%,():])
  | (?P<ws>\s+)
""", re.X)


def tokenize(f):
    toks, i = [], 0
    while i < len(f):
        m = TOKEN_RE.match(f, i)
        if not m:
            raise ValueError(f"tokenize failed at {i}: {f[i:i+30]!r} in {f!r}")
        i = m.end()
        kind = m.lastgroup
        if kind == 'ws':
            continue
        toks.append((kind, m.group()))
    return toks


# ---------------------------------------------------------------- parser
class Parser:
    def __init__(self, toks, cursheet):
        self.t, self.i, self.cur = toks, 0, cursheet

    def peek(self):
        return self.t[self.i] if self.i < len(self.t) else (None, None)

    def next(self):
        tok = self.t[self.i]; self.i += 1; return tok

    def expect(self, val):
        k, v = self.next()
        if v != val:
            raise ValueError(f"expected {val} got {v}")

    def parse(self):
        n = self.expr()
        if self.i != len(self.t):
            raise ValueError(f"trailing tokens {self.t[self.i:]}")
        return n

    def expr(self):
        n = self.concat()
        while self.peek()[1] in ('=', '<>', '<', '>', '<=', '>='):
            op = self.next()[1]
            n = ('cmp', op, n, self.concat())
        return n

    def concat(self):
        n = self.addsub()
        while self.peek()[1] == '&':
            self.next()
            n = ('cat', n, self.addsub())
        return n

    def addsub(self):
        n = self.muldiv()
        while self.peek()[1] in ('+', '-'):
            op = self.next()[1]
            n = ('bin', op, n, self.muldiv())
        return n

    def muldiv(self):
        n = self.power()
        while self.peek()[1] in ('*', '/'):
            op = self.next()[1]
            n = ('bin', op, n, self.power())
        return n

    def power(self):
        n = self.unary()
        while self.peek()[1] == '^':
            self.next()
            n = ('bin', '^', n, self.unary())
        return n

    def unary(self):
        k, v = self.peek()
        if v == '-':
            self.next(); return ('neg', self.unary())
        if v == '+':
            self.next(); return self.unary()
        return self.postfix()

    def postfix(self):
        n = self.atom()
        while self.peek()[1] == '%':
            self.next(); n = ('pct', n)
        return n

    def atom(self):
        k, v = self.next()
        if k == 'num':
            return ('num', float(v))
        if k == 'str':
            return ('str', v[1:-1].replace('""', '"'))
        if k == 'err':
            return ('err', v)
        if k == 'func':
            name = v[:-1].upper()
            args = []
            if self.peek()[1] != ')':
                args.append(self.expr())
                while self.peek()[1] == ',':
                    self.next(); args.append(self.expr())
            self.expect(')')
            return ('call', name, args)
        if v == '(':
            n = self.expr(); self.expect(')'); return n
        if k == 'sheet':
            sh = v[:-1]
            if sh.startswith("'"):
                sh = sh[1:-1].replace("''", "'")
            k2, v2 = self.next()
            return self.make_ref(sh, k2, v2)
        if k in ('ref', 'colrange'):
            return self.make_ref(self.cur, k, v)
        raise ValueError(f"unexpected token {k} {v}")

    def make_ref(self, sheet, k, v):
        if k == 'colrange':
            a, b = v.split(':')
            return ('colrange', sheet, a.replace('$', ''), b.replace('$', ''))
        if k != 'ref':
            raise ValueError(f"bad ref token {k} {v}")
        first = v.replace('$', '')
        if self.peek()[1] == ':':
            save = self.i
            self.next()
            k2, v2 = self.peek()
            if k2 == 'ref':
                self.next()
                return ('range', sheet, first, v2.replace('$', ''))
            self.i = save
        return ('cell', sheet, first)


# ---------------------------------------------------------------- engine
class Book:
    def __init__(self, path):
        import xlread
        self.formulas, self.values = xlread.read(path)
        self.unresolved = self.formulas.pop('__unresolved__', [])
        self.cache = {}          # (sheet, coord) -> value
        self.ast = {}
        self.stack = []
        self.cycles = set()
        self.raw = {}
        self.maxrow = {}
        for sheet, vals in self.values.items():
            d = dict(vals)
            for coord, f in self.formulas.get(sheet, {}).items():
                d[coord] = '=' + f
            self.raw[sheet] = d
            mr = 1
            for coord in d:
                r = int(re.match(r'[A-Z]+(\d+)', coord).group(1))
                if r > mr: mr = r
            self.maxrow[sheet] = mr
        self.sheetnames = list(self.values.keys())

    # --- value access -----------------------------------------------------
    def get(self, sheet, coord):
        key = (sheet, coord)
        if key in self.cache:
            return self.cache[key]
        if key in self.stack:
            self.cycles.add(key)
            return 0.0
        v = self.raw.get(sheet, {}).get(coord)
        if v is None:
            self.cache[key] = BLANK
            return BLANK
        if isinstance(v, str) and v.startswith('='):
            self.stack.append(key)
            try:
                node = self.ast.get(key)
                if node is None:
                    try:
                        node = Parser(tokenize(v[1:]), sheet).parse()
                    except Exception as e:
                        node = ('err', '#NAME?')
                        self.parse_fail = getattr(self, 'parse_fail', [])
                        self.parse_fail.append((sheet, coord, v, str(e)))
                    self.ast[key] = node
                out = self.eval(node, sheet)
            finally:
                self.stack.pop()
        elif isinstance(v, str) and v in ERRORS:
            out = Err(v)
        else:
            out = v
        self.cache[key] = out
        return out

    def cells_in_range(self, sheet, a, b):
        c1 = column_index_from_string(re.match(r'([A-Z]+)', a).group(1))
        r1 = int(re.match(r'[A-Z]+(\d+)', a).group(1))
        c2 = column_index_from_string(re.match(r'([A-Z]+)', b).group(1))
        r2 = int(re.match(r'[A-Z]+(\d+)', b).group(1))
        if c1 > c2: c1, c2 = c2, c1
        if r1 > r2: r1, r2 = r2, r1
        return [(r, c) for r in range(r1, r2 + 1) for c in range(c1, c2 + 1)]

    def range_values(self, node):
        kind = node[0]
        if kind == 'cell':
            return [self.get(node[1], node[2])]
        if kind == 'range':
            sh = node[1]
            return [self.get(sh, f"{get_column_letter(c)}{r}")
                    for r, c in self.cells_in_range(sh, node[2], node[3])]
        if kind == 'colrange':
            sh = node[1]
            c1 = column_index_from_string(node[2]); c2 = column_index_from_string(node[3])
            if c1 > c2: c1, c2 = c2, c1
            n = self.maxrow.get(sh, 1)
            return [self.get(sh, f"{get_column_letter(c)}{r}")
                    for r in range(1, n + 1) for c in range(c1, c2 + 1)]
        return [self.eval(node, None)]

    def range_shape(self, node):
        """(rows, cols, list-of-(sheet,row,col)) for SUMIF offsetting."""
        if node[0] == 'cell':
            m = re.match(r'([A-Z]+)(\d+)', node[2])
            c = column_index_from_string(m.group(1)); r = int(m.group(2))
            return node[1], r, c, 1, 1
        if node[0] == 'range':
            sh = node[1]
            m1 = re.match(r'([A-Z]+)(\d+)', node[2]); m2 = re.match(r'([A-Z]+)(\d+)', node[3])
            c1 = column_index_from_string(m1.group(1)); r1 = int(m1.group(2))
            c2 = column_index_from_string(m2.group(1)); r2 = int(m2.group(2))
            if c1 > c2: c1, c2 = c2, c1
            if r1 > r2: r1, r2 = r2, r1
            return sh, r1, c1, r2 - r1 + 1, c2 - c1 + 1
        if node[0] == 'colrange':
            sh = node[1]
            c1 = column_index_from_string(node[2]); c2 = column_index_from_string(node[3])
            if c1 > c2: c1, c2 = c2, c1
            n = self.maxrow.get(sh, 1)
            return sh, 1, c1, n, c2 - c1 + 1
        raise ValueError(f"not a range: {node[0]}")

    # --- evaluation -------------------------------------------------------
    def eval(self, n, sheet):
        k = n[0]
        if k == 'num':
            return n[1]
        if k == 'str':
            return n[1]
        if k == 'err':
            return Err(n[1])
        if k == 'cell':
            return self.get(n[1], n[2])
        if k in ('range', 'colrange'):
            vals = self.range_values(n)
            for v in vals:
                if isinstance(v, Err):
                    return v
            return vals[0] if vals else 0.0
        if k == 'neg':
            v = self.num(self.eval(n[1], sheet))
            return v if isinstance(v, Err) else -v
        if k == 'pct':
            v = self.num(self.eval(n[1], sheet))
            return v if isinstance(v, Err) else v / 100.0
        if k == 'cat':
            a = self.eval(n[1], sheet); b = self.eval(n[2], sheet)
            if isinstance(a, Err): return a
            if isinstance(b, Err): return b
            return self.txt(a) + self.txt(b)
        if k == 'cmp':
            a = self.eval(n[2], sheet); b = self.eval(n[3], sheet)
            if isinstance(a, Err): return a
            if isinstance(b, Err): return b
            if isinstance(a, str) or isinstance(b, str):
                if isinstance(a, Blank):
                    a = ''
                elif not isinstance(a, str):
                    a = self.txt(a)
                if isinstance(b, Blank):
                    b = ''
                elif not isinstance(b, str):
                    b = self.txt(b)
                a, b = a.upper(), b.upper()
            else:
                a, b = self.num(a), self.num(b)
                if isinstance(a, Err): return a
                if isinstance(b, Err): return b
            op = n[1]
            return {'=': a == b, '<>': a != b, '<': a < b, '>': a > b,
                    '<=': a <= b, '>=': a >= b}[op]
        if k == 'bin':
            a = self.num(self.eval(n[2], sheet))
            if isinstance(a, Err): return a
            b = self.num(self.eval(n[3], sheet))
            if isinstance(b, Err): return b
            op = n[1]
            if op == '+': return a + b
            if op == '-': return a - b
            if op == '*': return a * b
            if op == '/':
                return Err('#DIV/0!') if b == 0 else a / b
            if op == '^':
                try: return a ** b
                except Exception: return Err('#NUM!')
        if k == 'call':
            return self.call(n[1], n[2], sheet)
        raise ValueError(f"eval: unknown node {k}")

    # --- coercion ---------------------------------------------------------
    def num(self, v):
        if isinstance(v, Err): return v
        if v is None: return 0.0
        if isinstance(v, bool): return 1.0 if v else 0.0
        if isinstance(v, (int, float)): return float(v)
        if isinstance(v, datetime.datetime):
            return (v - datetime.datetime(1899, 12, 30)).days + \
                   (v.hour * 3600 + v.minute * 60 + v.second) / 86400.0
        if isinstance(v, datetime.date):
            return float((v - datetime.date(1899, 12, 30)).days)
        if isinstance(v, str):
            if v in ERRORS: return Err(v)
            s = v.strip().replace(',', '')
            if s.endswith('%'):
                try: return float(s[:-1]) / 100.0
                except ValueError: return Err('#VALUE!')
            try: return float(s)
            except ValueError: return Err('#VALUE!')
        return Err('#VALUE!')

    def txt(self, v):
        if isinstance(v, bool): return 'TRUE' if v else 'FALSE'
        if isinstance(v, float) and v == int(v): return str(int(v))
        if v is None: return ''
        return str(v)

    # --- functions --------------------------------------------------------
    def call(self, name, args, sheet):
        if name == 'IFERROR':
            v = self.eval(args[0], sheet)
            return self.eval(args[1], sheet) if isinstance(v, Err) else v
        if name == 'IF':
            c = self.eval(args[0], sheet)
            if isinstance(c, Err): return c
            c = bool(self.num(c)) if not isinstance(c, bool) else c
            if c:
                return self.eval(args[1], sheet) if len(args) > 1 else True
            return self.eval(args[2], sheet) if len(args) > 2 else False
        if name == 'TRUE':
            return True
        if name == 'FALSE':
            return False
        if name in ('SUM', 'MAX', 'MIN', 'AVERAGE', 'COUNT', 'PRODUCT'):
            vals = []
            for a in args:
                if a[0] in ('cell', 'range', 'colrange'):
                    vals.extend(self.range_values(a))
                else:
                    vals.append(self.eval(a, sheet))
            nums = []
            for v in vals:
                if isinstance(v, Err): return v
                if isinstance(v, bool): continue
                if isinstance(v, (int, float)): nums.append(float(v))
                elif isinstance(v, (datetime.date, datetime.datetime)):
                    nums.append(self.num(v))
            if name == 'SUM': return math.fsum(nums)
            if name == 'COUNT': return float(len(nums))
            if not nums: return 0.0
            if name == 'MAX': return max(nums)
            if name == 'MIN': return min(nums)
            if name == 'AVERAGE': return math.fsum(nums) / len(nums)
            p = 1.0
            for x in nums: p *= x
            return p
        if name in ('AND', 'OR'):
            vals = []
            for a in args:
                if a[0] in ('cell', 'range', 'colrange'):
                    vals.extend(self.range_values(a))
                else:
                    vals.append(self.eval(a, sheet))
            bs = []
            for v in vals:
                if isinstance(v, Err): return v
                if isinstance(v, bool): bs.append(v)
                elif isinstance(v, (int, float)): bs.append(v != 0)
                elif v is None or v == '': continue
                else: bs.append(bool(v))
            if not bs: return Err('#VALUE!')
            return all(bs) if name == 'AND' else any(bs)
        if name == 'NOT':
            v = self.eval(args[0], sheet)
            return v if isinstance(v, Err) else not bool(self.num(v))
        if name == 'ISNUMBER':
            v = self.eval(args[0], sheet)
            return isinstance(v, (int, float)) and not isinstance(v, bool)
        if name == 'ISERROR':
            return isinstance(self.eval(args[0], sheet), Err)
        if name == 'ISBLANK':
            v = self.eval(args[0], sheet)
            return v is None or v == '' or v == 0.0
        if name == 'SEARCH':
            find = self.eval(args[0], sheet); within = self.eval(args[1], sheet)
            if isinstance(find, Err): return find
            if isinstance(within, Err): return within
            i = self.txt(within).upper().find(self.txt(find).upper())
            return Err('#VALUE!') if i < 0 else float(i + 1)
        if name == 'MONTH':
            v = self.eval(args[0], sheet)
            if isinstance(v, Err): return v
            if isinstance(v, (datetime.date, datetime.datetime)): return float(v.month)
            n = self.num(v)
            if isinstance(n, Err): return n
            d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=n)
            return float(d.month)
        if name == 'YEAR':
            v = self.eval(args[0], sheet)
            if isinstance(v, Err): return v
            if isinstance(v, (datetime.date, datetime.datetime)): return float(v.year)
            d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=self.num(v))
            return float(d.year)
        if name == 'ROUND':
            v = self.num(self.eval(args[0], sheet))
            if isinstance(v, Err): return v
            d = int(self.num(self.eval(args[1], sheet))) if len(args) > 1 else 0
            return round(v, d)
        if name == 'ABS':
            v = self.num(self.eval(args[0], sheet))
            return v if isinstance(v, Err) else abs(v)
        if name in ('SUMIF', 'COUNTIF', 'AVERAGEIF'):
            return self.sumif(name, args, sheet)
        if name == 'SUMPRODUCT':
            arrays = [self.range_values(a) if a[0] in ('cell', 'range', 'colrange')
                      else [self.eval(a, sheet)] for a in args]
            n = max(len(a) for a in arrays)
            tot = 0.0
            for i in range(n):
                p = 1.0
                for a in arrays:
                    v = a[i] if len(a) > 1 else a[0]
                    if isinstance(v, Err): return v
                    nv = self.num(v)
                    p *= 0.0 if isinstance(nv, Err) else nv
                tot += p
            return tot
        if name == 'CONCAT' or name == 'CONCATENATE':
            out = []
            for a in args:
                vs = self.range_values(a) if a[0] in ('range', 'colrange') else [self.eval(a, sheet)]
                for v in vs:
                    if isinstance(v, Err): return v
                    out.append(self.txt(v))
            return ''.join(out)
        if name == 'UPPER':
            v = self.eval(args[0], sheet)
            return v if isinstance(v, Err) else self.txt(v).upper()
        return Err('#NAME?')

    def sumif(self, name, args, sheet):
        rsheet, r0, c0, nr, nc = self.range_shape(args[0])
        crit = self.eval(args[1], sheet)
        if isinstance(crit, Err): return crit
        if len(args) > 2:
            ssheet, sr0, sc0, _, _ = self.range_shape(args[2])
        else:
            ssheet, sr0, sc0 = rsheet, r0, c0
        op, target = '=', crit
        if isinstance(crit, str):
            m = re.match(r'^(<=|>=|<>|<|>)(.*)$', crit)
            if m:
                op, target = m.group(1), m.group(2)
                try: target = float(target)
                except ValueError: pass
        tot, cnt = 0.0, 0
        for i in range(nr):
            for j in range(nc):
                cv = self.get(rsheet, f"{get_column_letter(c0+j)}{r0+i}")
                ok = self.match(cv, op, target)
                if not ok:
                    continue
                if name == 'COUNTIF':
                    cnt += 1; continue
                sv = self.get(ssheet, f"{get_column_letter(sc0+j)}{sr0+i}")
                if isinstance(sv, Err):
                    return sv
                nv = self.num(sv)
                if not isinstance(nv, Err):
                    tot += nv; cnt += 1
        if name == 'COUNTIF': return float(cnt)
        if name == 'AVERAGEIF': return tot / cnt if cnt else Err('#DIV/0!')
        return tot

    @staticmethod
    def match(cv, op, target):
        if isinstance(cv, Err):
            return False
        if op == '=':
            if isinstance(target, str):
                if cv is None: return target == ''
                if isinstance(cv, str): return cv.strip().upper() == target.strip().upper()
                return False
            if isinstance(cv, (int, float)) and not isinstance(cv, bool):
                return float(cv) == float(target)
            if cv is None:
                return float(target) == 0.0 and False
            return False
        if isinstance(target, str) or not isinstance(cv, (int, float)) or isinstance(cv, bool):
            return False
        a, b = float(cv), float(target)
        return {'<': a < b, '>': a > b, '<=': a <= b, '>=': a >= b, '<>': a != b}[op]


def load(path):
    return Book(path)

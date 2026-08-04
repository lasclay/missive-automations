#!/usr/bin/env python3
"""Phase AS — passer juillet 2026 du prévisionnel au réel.

Un mois réel et un mois prévisionnel ne se distinguent pas par une case à
cocher : ils se distinguent par la **forme des formules**. Un mois réel va
chercher QuickBooks (`SUMIF` sur les feuilles de collage) ; un mois
prévisionnel applique les moteurs du modèle (Immos, Dette à long terme,
Account payable…). Basculer juillet, c'est donc réécrire sa colonne pour
qu'elle ait exactement la forme de celle de juin.

## Pourquoi une transposition intégrale, et pas un tri rangée par rangée

Un premier essai avait basculé les seules rangées où QuickBooks porte un
montant en juillet, en laissant au modèle les rangées qu'il calcule lui-même —
amortissements, intérêts, chaîne de trésorerie. Le résultat ne tombait juste
d'aucun côté : ni le réel de QuickBooks, ni la prévision cohérente. Un mois
mi-réel mi-prévu n'est **pas** un compromis prudent, c'est un mois faux.

La forme correcte est celle de juin, sans exception. Les amortissements
tombent à zéro parce que QuickBooks ne les passe pas au mois — c'est déjà vrai
des dix autres mois réels de l'exercice, et le nier pour juillet seul créerait
l'incohérence qu'on cherche à éviter.

## La transposition elle-même

`xlread.translate` fait ce que fait Excel quand on tire une formule d'une
colonne vers la droite : les références relatives se décalent, les ancrées ne
bougent pas. `M9 → N9` transforme donc du même geste `'QBO P&L à maj'!AJ:AJ` en
`AK:AK` (juin → juillet dans la feuille de collage), `M27` en `N27`, et laisse
`$B9` intact. Écrire ces décalages à la main serait dix occasions de se
tromper.

## Le bilan doit suivre le résultat

La colonne de trésorerie du résultat lit le bilan (`N189 = Bilan!N6`), et le
bilan réel lit QuickBooks. Basculer le résultat sans le bilan ferait lire une
encaisse prévisionnelle à un mois réel. Les deux feuilles basculent ensemble.

    python3 fix_as.py --essai
    python3 fix_as.py
"""
import re
import sys

sys.path.insert(0, 'tools')
from openpyxl.utils import column_index_from_string, get_column_letter
from xledit import Editor
import xlcalc
import xlread

F = 'PREVISIONS LASCLAY - version audit 2026-07-30.xlsx'

# `xlread.translate` ne décale que les références qui portent un numéro de
# rangée. Une colonne entière — `'QBO P&L à maj'!AJ:AJ`, la forme qu'emploient
# les `SUMIF` du résultat — n'en porte pas : elle traverse la transposition
# intacte, et juillet se met à lire la colonne de juin. Le premier passage a
# ainsi recopié juin dans juillet, au dollar près, sans qu'aucune formule ne
# soit visiblement fausse. C'est le genre d'erreur qui se voit seulement si on
# compare le résultat à la source.
PLAGE_COL = re.compile(r'(\$?)([A-Z]{1,3}):(\$?)([A-Z]{1,3})(?![A-Z0-9])')


def decaler_colonnes(formule, n):
    """Décale de `n` les plages de colonnes entières non ancrées."""
    out, i = [], 0
    while i < len(formule):
        m = xlread.SKIP_RE.match(formule, i)     # noms de feuille, littéraux
        if m:
            out.append(formule[i:m.end()])
            i = m.end()
            continue
        m = PLAGE_COL.match(formule, i)
        if m:
            d1, c1, d2, c2 = m.groups()
            if not d1:
                c1 = get_column_letter(column_index_from_string(c1) + n)
            if not d2:
                c2 = get_column_letter(column_index_from_string(c2) + n)
            out.append(f'{d1}{c1}:{d2}{c2}')
            i = m.end()
            continue
        out.append(formule[i])
        i += 1
    return ''.join(out)

# feuille, colonne du mois réel de référence, colonne à basculer
BASCULE = [
    ('Résultats-Prev 2025-2029', 'M', 'N'),
    ('Bilan2026-27-28', 'M', 'N'),
]


def build(dst=F, src=F, essai=False):
    formules, _ = xlread.read(src)
    avant = xlcalc.load(src)
    e = Editor(src)
    e.expand_shared()
    resume = []

    for feuille, ref, cible in BASCULE:
        f = formules[feuille]
        n = column_index_from_string(cible) - column_index_from_string(ref)
        rangees = sorted({int(k[len(ref):]) for k in f
                          if k.startswith(ref) and k[len(ref):].isdigit()})
        ecrites = identiques = 0
        for r in rangees:
            source = f[f'{ref}{r}']
            neuve = xlread.translate(decaler_colonnes(source, n),
                                     f'{ref}{r}', f'{cible}{r}')
            if f.get(f'{cible}{r}') == neuve:
                identiques += 1
                continue
            if not essai:
                e.set(feuille, f'{cible}{r}', '=' + neuve)
            ecrites += 1
        # L'en-tête dit au lecteur ce que la forme des formules dit déjà au
        # classeur ; les deux doivent concorder, sinon le lecteur se fie au
        # mauvais des deux.
        if not essai:
            e.set(feuille, f'{cible}2', 'Actuel')
        resume.append(f'{feuille} : {ecrites} formules transposées de '
                      f'{ref} vers {cible}, {identiques} déjà identiques, '
                      f'en-tête « Forecast » → « Actuel »')

    if essai:
        return resume

    e.set_full_calc()
    e.save(dst)

    apres = xlcalc.load(dst)
    S = 'Résultats-Prev 2025-2029'
    for lib, ref in (('Ventes nettes', 'N26'), ('Résultat avant impôts', 'N137'),
                     ('Encaisse à la fin', 'N189'),
                     ('Marge de crédit tirée', 'N183')):
        resume.append(f'  {lib:<24} {avant.get(S, ref):>14,.2f} → '
                      f'{apres.get(S, ref):>14,.2f}')
    resume.append(f"  contrôle du bilan (rangée 76) : "
                  f"{apres.get('Bilan2026-27-28', 'N76'):.6f}")
    return resume


if __name__ == '__main__':
    for ligne in build(essai='--essai' in sys.argv):
        print(ligne)
